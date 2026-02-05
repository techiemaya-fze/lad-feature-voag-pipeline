"""
Combined Batch Report Generator

Fetches call data for multiple batch IDs and generates a single combined Excel report.
"""
import asyncio
import os
import sys
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Dict

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv()

# Import from batch_report module
from analysis.batch_report import (
    get_batch_info,
    fetch_batch_call_data,
    _convert_to_gst,
    _get_time_period,
    _format_duration,
    _format_list_field,
    _apply_enhanced_excel_formatting,
    EXCEL_AVAILABLE,
    GST,
)

if not EXCEL_AVAILABLE:
    print("ERROR: pandas/openpyxl not installed. Install with: pip install pandas openpyxl")
    sys.exit(1)

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _format_summary_sheet(worksheet, df, batch_infos: List[Dict], all_data: List[Dict]):
    """Apply professional executive-style formatting to the Summary sheet"""
    
    # Professional color palette
    HEADER_DARK = "0F172A"
    ACCENT_BLUE = "3B82F6"
    BORDER_LIGHT = "E2E8F0"
    TEXT_DARK = "0F172A"
    ALT_ROW_LIGHT = "F8FAFC"
    STATUS_SUCCESS = "D1FAE5"
    STATUS_SUCCESS_TEXT = "065F46"
    
    # Clear existing content and rebuild with better layout
    # Delete all rows first
    worksheet.delete_rows(1, worksheet.max_row)
    
    # Row counter
    row = 1
    
    # Title section
    worksheet.merge_cells(f'A{row}:F{row}')
    title_cell = worksheet[f'A{row}']
    title_cell.value = "VOICE AGENT ANALYTICS PLATFORM"
    title_cell.font = Font(bold=True, size=10, color=ACCENT_BLUE, name="Inter")
    title_cell.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[row].height = 20
    row += 1
    
    # Main heading
    worksheet.merge_cells(f'A{row}:F{row}')
    main_title = worksheet[f'A{row}']
    main_title.value = "Combined Batch Report - Executive Summary"
    main_title.font = Font(bold=True, size=20, color="FFFFFF", name="Inter")
    main_title.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    main_title.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[row].height = 40
    row += 1
    
    # Subtitle with timestamp
    worksheet.merge_cells(f'A{row}:F{row}')
    subtitle = worksheet[f'A{row}']
    subtitle.value = f"Generated {datetime.now(GST).strftime('%B %d, %Y at %H:%M GST')} • {len(batch_infos)} Batches Combined"
    subtitle.font = Font(size=11, color="CBD5E1", name="Inter")
    subtitle.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    subtitle.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[row].height = 28
    row += 2  # Extra space
    
    # Key Metrics Section Header
    worksheet.merge_cells(f'A{row}:F{row}')
    metrics_header = worksheet[f'A{row}']
    metrics_header.value = "KEY METRICS"
    metrics_header.font = Font(bold=True, size=14, color=TEXT_DARK, name="Inter")
    metrics_header.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[row].height = 30
    row += 1
    
    # Calculate totals
    total_calls = sum(b.get('total_calls', 0) for b in batch_infos)
    completed_calls = sum(b.get('completed_calls', 0) for b in batch_infos)
    failed_calls = sum(b.get('failed_calls', 0) for b in batch_infos)
    calls_with_analysis = len([c for c in all_data if c.get('summary')])
    success_rate = round((completed_calls / total_calls * 100), 1) if total_calls > 0 else 0
    analysis_rate = round((calls_with_analysis / len(all_data) * 100), 1) if all_data else 0
    
    # Metrics cards in a row
    metrics = [
        ("Total Batches", str(len(batch_infos)), "F1F5F9", TEXT_DARK),
        ("Total Calls", str(total_calls), "F1F5F9", TEXT_DARK),
        ("Completed", str(completed_calls), "D1FAE5", "065F46"),
        ("Failed", str(failed_calls), "FEE2E2", "991B1B"),
        ("With Analysis", str(calls_with_analysis), "DBEAFE", "1E40AF"),
        ("Success Rate", f"{success_rate}%", "D1FAE5", "065F46"),
    ]
    
    col = 1
    for label, value, bg_color, text_color in metrics:
        cell = worksheet.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(bold=True, size=24, color=text_color, name="Inter")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color=BORDER_LIGHT),
            right=Side(style='thin', color=BORDER_LIGHT),
            top=Side(style='thin', color=BORDER_LIGHT),
            bottom=Side(style='thin', color=BORDER_LIGHT)
        )
        col += 1
    
    worksheet.row_dimensions[row].height = 50
    row += 1
    
    # Metric labels
    col = 1
    for label, value, bg_color, text_color in metrics:
        cell = worksheet.cell(row=row, column=col)
        cell.value = label
        cell.font = Font(size=10, color="64748B", name="Inter")
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color=BORDER_LIGHT),
            right=Side(style='thin', color=BORDER_LIGHT),
            top=Side(style='thin', color=BORDER_LIGHT),
            bottom=Side(style='thin', color=BORDER_LIGHT)
        )
        col += 1
    
    worksheet.row_dimensions[row].height = 25
    row += 2  # Extra space
    
    # Batch Details Section Header
    worksheet.merge_cells(f'A{row}:F{row}')
    batch_header = worksheet[f'A{row}']
    batch_header.value = "BATCH DETAILS"
    batch_header.font = Font(bold=True, size=14, color=TEXT_DARK, name="Inter")
    batch_header.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[row].height = 30
    row += 1
    
    # Batch table headers
    headers = ["#", "Batch ID", "Status", "Total", "Completed", "Failed"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color="475569", name="Inter")
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color=BORDER_LIGHT),
            right=Side(style='thin', color=BORDER_LIGHT),
            top=Side(style='medium', color=BORDER_LIGHT),
            bottom=Side(style='medium', color=BORDER_LIGHT)
        )
    worksheet.row_dimensions[row].height = 30
    row += 1
    
    # Batch rows
    for idx, batch_info in enumerate(batch_infos, 1):
        is_even = idx % 2 == 0
        row_bg = ALT_ROW_LIGHT if is_even else "FFFFFF"
        
        values = [
            idx,
            batch_info.get('id', 'N/A'),
            batch_info.get('status', 'N/A').title(),
            batch_info.get('total_calls', 0),
            batch_info.get('completed_calls', 0),
            batch_info.get('failed_calls', 0),
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = Font(size=10, color="334155", name="Inter")
            cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='hair', color=BORDER_LIGHT),
                right=Side(style='hair', color=BORDER_LIGHT),
                top=Side(style='hair', color="F1F5F9"),
                bottom=Side(style='thin', color="F1F5F9")
            )
            
            # Special formatting for status column
            if col_idx == 3:
                status = str(value).lower()
                if status == 'completed':
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    cell.font = Font(bold=True, size=10, color="065F46", name="Inter")
                elif status == 'running':
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    cell.font = Font(bold=True, size=10, color="1E40AF", name="Inter")
                elif status in ['failed', 'error']:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(bold=True, size=10, color="991B1B", name="Inter")
        
        worksheet.row_dimensions[row].height = 28
        row += 1
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12

# Batch IDs from user request
BATCH_IDS = [
    "0f85e856-3b26-4c73-9d9b-1080f009fef4",  # 75 calls
    "62c2dcb1-15eb-4379-bdae-ac74238f464f",  # 75 calls
    "08641b4f-7ce1-44ab-90a8-d3f5c2c4e81a",  # 75 calls
    "64a2034f-e1ad-45ff-a4e5-f04ffd3aac97",  # 75 calls
    "8de7df76-599b-4751-8cc1-060e033ac767",  # 42 calls
]


async def fetch_all_batches_data() -> tuple[List[Dict], List[Dict]]:
    """Fetch call data for all batches and return combined data plus batch infos"""
    all_data = []
    batch_infos = []
    
    for batch_id in BATCH_IDS:
        print(f"\n--- Fetching batch: {batch_id} ---")
        
        # Get batch info
        batch_info = await get_batch_info(batch_id)
        if batch_info:
            batch_infos.append(batch_info)
            print(f"  Status: {batch_info['status']}, Total: {batch_info['total_calls']}, "
                  f"Completed: {batch_info['completed_calls']}, Failed: {batch_info['failed_calls']}")
        else:
            print(f"  WARNING: Batch {batch_id} not found!")
            continue
        
        # Fetch call data
        call_data = await fetch_batch_call_data(batch_id)
        print(f"  Fetched {len(call_data)} call records")
        
        # Add batch_id to each record for reference
        for record in call_data:
            record['batch_id'] = batch_id
        
        all_data.extend(call_data)
    
    return all_data, batch_infos


def export_combined_excel(
    data: List[Dict],
    batch_infos: List[Dict],
    output_file: str = "combined_batch_report.xlsx"
) -> str:
    """Export combined batch data to Excel file in project root"""
    
    if not data:
        raise ValueError("No data to export")
    
    output_path = Path(__file__).parent / output_file
    
    # Prepare data for Excel
    excel_data = []
    for record in data:
        started_at = record.get('started_at')
        ended_at = record.get('ended_at')
        
        # Convert timestamps to Gulf Standard Time (UTC+4)
        gst_started_at = _convert_to_gst(started_at) if isinstance(started_at, datetime) else None
        gst_ended_at = _convert_to_gst(ended_at) if isinstance(ended_at, datetime) else None
        
        # Format timestamps in GST
        date_str = gst_started_at.strftime('%Y-%m-%d') if gst_started_at else 'N/A'
        start_time_str = gst_started_at.strftime('%H:%M:%S') if gst_started_at else 'N/A'
        end_time_str = gst_ended_at.strftime('%H:%M:%S') if gst_ended_at else 'N/A'
        
        row = {
            'Batch ID': str(record.get('batch_id', ''))[:8] + '...',  # Shortened for readability
            'Call Log ID': str(record.get('call_log_id', '')) if record.get('call_log_id') else '',
            'Lead Name': record.get('lead_name', '') or '',
            'Lead Number': record.get('lead_number', '') or '',
            'Date': date_str,
            'Start Time (GST)': start_time_str,
            'End Time (GST)': end_time_str,
            'Duration': _format_duration(started_at, ended_at),
            'Call Status': record.get('call_status', '') or '',
            'Time Period (GST)': _get_time_period(started_at),
            'Disposition': record.get('disposition', '') or '',
            'Recommended Action': record.get('recommended_action', '') or '',
            'Lead Category': record.get('lead_category', '') or '',
            'Engagement Level': record.get('engagement_level', '') or '',
            'Sentiment': record.get('sentiment_description', '') or '',
            'Summary': record.get('summary', '') or '',
            'Key Discussion Points': _format_list_field(record.get('key_discussion_points')),
            'Prospect Questions': _format_list_field(record.get('prospect_questions')),
            'Prospect Concerns': _format_list_field(record.get('prospect_concerns')),
            'Recommendations': record.get('recommendations', '') or '',
        }
        excel_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Column order
    column_order = [
        'Batch ID', 'Call Log ID', 'Lead Name', 'Lead Number', 'Date', 'Start Time (GST)',
        'End Time (GST)', 'Duration', 'Call Status', 'Time Period (GST)', 'Disposition',
        'Recommended Action', 'Lead Category', 'Engagement Level', 'Sentiment',
        'Summary', 'Key Discussion Points', 'Prospect Questions',
        'Prospect Concerns', 'Recommendations'
    ]
    
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary sheet with batch info
        summary_data = {
            'Metric': ['Total Batches', 'Total Calls', 'Calls with Analysis', 'Report Generated'],
            'Value': [
                len(batch_infos),
                len(data),
                len([c for c in data if c.get('summary')]),
                datetime.now(GST).strftime('%Y-%m-%d %H:%M:%S GST'),
            ]
        }
        
        # Add per-batch info
        for i, batch_info in enumerate(batch_infos, 1):
            summary_data['Metric'].extend([
                f'--- Batch {i} ---',
                f'Batch ID',
                f'Status',
                f'Total',
                f'Completed',
                f'Failed',
            ])
            summary_data['Value'].extend([
                '',
                batch_info.get('id', 'N/A'),
                batch_info.get('status', 'N/A'),
                batch_info.get('total_calls', 0),
                batch_info.get('completed_calls', 0),
                batch_info.get('failed_calls', 0),
            ])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply professional formatting to Summary sheet
        _format_summary_sheet(writer.sheets['Summary'], summary_df, batch_infos, data)
        
        # All calls sheet
        df.to_excel(writer, sheet_name='All Calls', index=False)
        _apply_enhanced_excel_formatting(writer.sheets['All Calls'], df)
        
        # Group by time period if multiple records
        if len(data) > 1 and 'Time Period (GST)' in df.columns:
            time_periods = df['Time Period (GST)'].unique()
            for period in sorted(time_periods):
                if pd.isna(period) or period == 'Unknown':
                    continue
                period_df = df[df['Time Period (GST)'] == period].copy()
                period_df = period_df.drop(columns=['Time Period (GST)'], errors='ignore')
                
                sheet_name = period[:31].replace(':', '-').replace('/', '-')
                period_df.to_excel(writer, sheet_name=sheet_name, index=False)
                try:
                    _apply_enhanced_excel_formatting(writer.sheets[sheet_name], period_df)
                except Exception as e:
                    print(f"  Warning: Could not format sheet {sheet_name}: {e}")
    
    print(f"\n✓ Excel report created: {output_path}")
    print(f"  Total records: {len(data)}")
    print(f"  With analysis: {len([c for c in data if c.get('summary')])}")
    
    return str(output_path)


async def main():
    print("=" * 60)
    print("COMBINED BATCH REPORT GENERATOR")
    print("=" * 60)
    print(f"\nGenerating combined report for {len(BATCH_IDS)} batches...")
    
    # Fetch all data
    all_data, batch_infos = await fetch_all_batches_data()
    
    print(f"\n{'=' * 60}")
    print(f"TOTAL: {len(all_data)} call records from {len(batch_infos)} batches")
    print(f"{'=' * 60}")
    
    if not all_data:
        print("\nERROR: No data found for any batch!")
        return
    
    # Export to Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"combined_batch_report_{timestamp}.xlsx"
    
    excel_path = export_combined_excel(all_data, batch_infos, output_file)
    
    print(f"\n{'=' * 60}")
    print("REPORT COMPLETE!")
    print(f"{'=' * 60}")
    print(f"Output: {excel_path}")


if __name__ == "__main__":
    asyncio.run(main())
