"""
Call Report Service - Standalone Service for Single and Bulk Call Reports
Handles both single call reports and bulk call reports based on batch_id

Phase 13: Updated to use dynamic schema (no longer uses voice_agent)
"""

import os
import json
import psycopg2
import asyncio
from datetime import datetime, timedelta
from dotenv import load_dotenv
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Optional, List, Dict
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

# Schema constants for table names
from db.schema_constants import (
    CALL_LOGS_FULL,
    ANALYSIS_FULL,
    LEADS_FULL,
    BATCHES_FULL,
    BATCH_ENTRIES_FULL,
)

load_dotenv()

# Database configuration
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", "5432")),
    "database": os.getenv("DB_NAME", "salesmaya_agent"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD"),
}

# Create directories for exports
EXPORTS_DIR = Path(__file__).parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


def get_db_connection():
    """Create database connection"""
    return psycopg2.connect(**DB_CONFIG)


def get_email_recipients(env_key: str = "REPORT_EMAIL") -> List[str]:
    """Get list of email recipients from environment variable"""
    emails_str = os.getenv(env_key, "").strip()
    if not emails_str:
        return []
    
    # Remove quotes if present
    emails_str = emails_str.strip('"\'')
    
    # Split by comma, semicolon, or newline, then clean up
    emails = []
    
    # Try comma first (most common)
    if ',' in emails_str:
        emails = [e.strip() for e in emails_str.split(',')]
    # Try semicolon
    elif ';' in emails_str:
        emails = [e.strip() for e in emails_str.split(';')]
    # Try newline
    elif '\n' in emails_str:
        emails = [e.strip() for e in emails_str.split('\n')]
    # If no separator found, treat as single email
    else:
        emails = [emails_str.strip()]
    
    # Filter out empty strings and validate emails (basic check)
    emails = [e for e in emails if e and '@' in e]
    
    return emails


def fetch_single_call_data(call_log_id) -> Optional[Dict]:
    """
    Fetch single call analysis data from database
    
    Args:
        call_log_id: Call log ID (UUID string or integer row number)
    
    Returns:
        Dictionary with call data or None if not found
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Try to parse as integer first, otherwise treat as UUID string
        try:
            call_log_id_int = int(call_log_id)
            call_log_id = call_log_id_int
        except (ValueError, TypeError):
            call_log_id = call_log_id
        
        if isinstance(call_log_id, int):
            # Integer ID: Use ROW_NUMBER() to find the Nth record
            cursor.execute(f"""
                SELECT cl.started_at, cl.ended_at, a.call_log_id, 
                       COALESCE(l.first_name || ' ' || l.last_name, '') as name,
                       COALESCE('+' || l.country_code || l.base_number::text, '') as lead_number,
                       a.summary, a.sentiment, a.key_points, a.lead_extraction, a.raw_analysis
                FROM (
                    SELECT 
                        id, started_at, ended_at, lead_id,
                        ROW_NUMBER() OVER (ORDER BY created_at) as row_num
                    FROM {CALL_LOGS_FULL}
                ) ranked
                INNER JOIN {ANALYSIS_FULL} a ON a.call_log_id = ranked.id
                INNER JOIN {CALL_LOGS_FULL} cl ON a.call_log_id = cl.id
                LEFT JOIN {LEADS_FULL} l ON cl.lead_id = l.id
                WHERE ranked.row_num = %s
            """, (call_log_id,))
        else:
            # UUID string: Try direct UUID match
            cursor.execute(f"""
                SELECT cl.started_at, cl.ended_at, a.call_log_id, 
                       COALESCE(l.first_name || ' ' || l.last_name, '') as name,
                       COALESCE('+' || l.country_code || l.base_number::text, '') as lead_number,
                       a.summary, a.sentiment, a.key_points, a.lead_extraction, a.raw_analysis
                FROM {ANALYSIS_FULL} a
                INNER JOIN {CALL_LOGS_FULL} cl ON a.call_log_id = cl.id
                LEFT JOIN {LEADS_FULL} l ON cl.lead_id = l.id
                WHERE a.call_log_id = %s::uuid
            """, (str(call_log_id),))
        
        columns = [desc[0] for desc in cursor.description]
        row = cursor.fetchone()
        
        if not row:
            return None
        
        # Convert to dictionary
        call_data = dict(zip(columns, row))
        
        # Extract fields from lead_extraction JSONB for backwards compatibility
        lead_extraction = call_data.get('lead_extraction') or {}
        if isinstance(lead_extraction, str):
            try:
                lead_extraction = json.loads(lead_extraction)
            except:
                lead_extraction = {}
        
        # Map new schema fields to old report format for compatibility
        call_data['lead_category'] = lead_extraction.get('lead_category', '')
        call_data['disposition'] = lead_extraction.get('disposition', '')
        call_data['recommended_action'] = lead_extraction.get('recommended_action', '')
        call_data['engagement_level'] = lead_extraction.get('engagement_level', '')
        
        # Extract key_points from JSONB
        key_points = call_data.get('key_points') or []
        if isinstance(key_points, str):
            try:
                key_points = json.loads(key_points)
            except:
                key_points = []
        call_data['key_discussion_points'] = key_points
        call_data['key_phrases'] = []  # Not stored separately anymore
        call_data['prospect_questions'] = []  # Extract from raw_analysis if needed
        call_data['prospect_concerns'] = []  # Extract from raw_analysis if needed
        call_data['recommendations'] = lead_extraction.get('recommendations', '')
        
        return call_data
        
    finally:
        cursor.close()
        conn.close()



def check_batch_id_exists(batch_id: str) -> Dict:
    """
    Check if batch_id exists and get statistics
    
    Args:
        batch_id: Batch ID (UUID) to check
    
    Returns:
        Dictionary with batch_id existence and call counts
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check total calls with this batch_id (via batch_entries table)
        cursor.execute(f"""
            SELECT COUNT(*) as total_calls
            FROM {BATCHES_FULL} b
            WHERE b.id = %s::uuid
        """, (str(batch_id),))
        
        batch_row = cursor.fetchone()
        if not batch_row or batch_row[0] == 0:
            # Batch doesn't exist - check entries table
            cursor.execute(f"""
                SELECT COUNT(*) as total_calls
                FROM {BATCH_ENTRIES_FULL}
                WHERE batch_id = %s::uuid
            """, (str(batch_id),))
            total_calls = cursor.fetchone()[0]
        else:
            # Get total_calls from batch record
            cursor.execute(f"""
                SELECT total_calls FROM {BATCHES_FULL} WHERE id = %s::uuid
            """, (str(batch_id),))
            total_calls = cursor.fetchone()[0] or 0
        
        # Check calls with analysis
        cursor.execute(f"""
            SELECT COUNT(DISTINCT be.call_log_id) as calls_with_analysis
            FROM {BATCH_ENTRIES_FULL} be
            INNER JOIN {ANALYSIS_FULL} a ON be.call_log_id = a.call_log_id
            WHERE be.batch_id = %s::uuid
        """, (str(batch_id),))
        
        calls_with_analysis = cursor.fetchone()[0]
        
        return {
            "exists": total_calls > 0,
            "total_calls": total_calls,
            "calls_with_analysis": calls_with_analysis,
            "calls_without_analysis": total_calls - calls_with_analysis
        }
    finally:
        cursor.close()
        conn.close()


def fetch_bulk_call_data_by_batch_id(batch_id) -> List[Dict]:
    """
    Fetch all calls for a specific batch_id from database
    
    Args:
        batch_id: Batch ID (UUID) from voice_call_batches table
    
    Returns:
        List of dictionaries with call data
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = f"""
            SELECT 
                cl.started_at,
                cl.ended_at,
                a.call_log_id,
                COALESCE(l.first_name || ' ' || l.last_name, '') as name,
                COALESCE('+' || l.country_code || l.base_number::text, '') as lead_number,
                a.summary,
                a.sentiment,
                a.key_points,
                a.lead_extraction,
                a.raw_analysis,
                be.batch_id
            FROM {ANALYSIS_FULL} a
            INNER JOIN {CALL_LOGS_FULL} cl ON a.call_log_id = cl.id
            LEFT JOIN {LEADS_FULL} l ON cl.lead_id = l.id
            INNER JOIN {BATCH_ENTRIES_FULL} be ON be.call_log_id = cl.id
            WHERE be.batch_id = %s::uuid
            ORDER BY cl.started_at ASC
        """
        
        cursor.execute(query, (str(batch_id),))
        
        # Get column names
        columns = [desc[0] for desc in cursor.description]
        
        # Fetch all results
        rows = cursor.fetchall()
        
        # Convert to list of dictionaries
        data = []
        for row in rows:
            row_dict = dict(zip(columns, row))
            
            # Extract fields from lead_extraction JSONB for backwards compatibility
            lead_extraction = row_dict.get('lead_extraction') or {}
            if isinstance(lead_extraction, str):
                try:
                    lead_extraction = json.loads(lead_extraction)
                except:
                    lead_extraction = {}
            
            # Map new schema fields to old report format for compatibility
            row_dict['lead_category'] = lead_extraction.get('lead_category', '')
            row_dict['disposition'] = lead_extraction.get('disposition', '')
            row_dict['recommended_action'] = lead_extraction.get('recommended_action', '')
            row_dict['engagement_level'] = lead_extraction.get('engagement_level', '')
            
            # Extract key_points from JSONB
            key_points = row_dict.get('key_points') or []
            if isinstance(key_points, str):
                try:
                    key_points = json.loads(key_points)
                except:
                    key_points = []
            row_dict['key_discussion_points'] = key_points
            row_dict['key_phrases'] = []
            row_dict['prospect_questions'] = []
            row_dict['prospect_concerns'] = []
            row_dict['recommendations'] = lead_extraction.get('recommendations', '')
            
            data.append(row_dict)
        
        return data
        
    finally:
        cursor.close()
        conn.close()


def get_time_period(started_at: datetime) -> str:
    """Categorize call into time period based on start time"""
    if not started_at:
        return "Unknown"
    
    hour = started_at.hour
    
    # Define time periods
    if 9 <= hour < 12:
        return "09:00-12:00 (Morning)"
    elif 12 <= hour < 14:
        return "12:00-14:00 (Midday)"
    elif 14 <= hour < 17:
        return "14:00-17:00 (Afternoon)"
    elif 17 <= hour < 20:
        return "17:00-20:00 (Evening)"
    elif 20 <= hour < 24:
        return "20:00-24:00 (Night)"
    else:
        return "00:00-09:00 (Early Morning)"


def format_json_field(value):
    """Format JSON field for Excel display"""
    if isinstance(value, dict):
        return json.dumps(value, indent=2)
    elif isinstance(value, list):
        return ", ".join(str(item) for item in value) if value else ""
    elif value is None:
        return ""
    else:
        return str(value)


def _format_worksheet(worksheet, df):
    """Apply professional formatting to Excel worksheet"""
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    center_alignment = Alignment(vertical='center', horizontal='center')
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 25
    
    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = wrap_alignment
            worksheet.row_dimensions[row_idx].height = 20
    
    # Set column widths
    column_widths = {
        'Call Log ID': 40,
        'Date': 12,
        'Start Time': 12,
        'End Time': 12,
        'Duration': 12,
        'Time Period': 20,
        'Disposition': 15,
        'Recommended Action': 40,
        'Lead Category': 15,
        'Engagement Level': 18,
        'Sentiment': 60,
        'Summary': 60,
        'Key Phrases': 40,
        'Key Discussion Points': 50,
        'Prospect Questions': 50,
        'Prospect Concerns': 50,
        'Recommendations': 50,
    }
    
    for idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        if col in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col]
        else:
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 60)
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Enable filter on header row
    worksheet.auto_filter.ref = worksheet.dimensions


def export_to_excel(data: List[Dict], output_file: str = None, batch_id: str = None) -> str:
    """
    Export analysis data to Excel file
    
    Args:
        data: List of analysis records
        output_file: Output Excel file path (optional)
        batch_id: Batch ID for filename (optional)
    
    Returns:
        Path to exported Excel file
    """
    if not data:
        raise ValueError("No data to export!")
    
    # Generate output filename if not provided
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if batch_id:
            output_file = f"call_analysis_batch_{batch_id}_{timestamp}.xlsx"
        elif len(data) == 1:
            call_id = str(data[0].get('call_log_id', 'single'))
            output_file = f"call_analysis_single_{call_id}_{timestamp}.xlsx"
        else:
            output_file = f"call_analysis_export_{timestamp}.xlsx"
    
    output_path = EXPORTS_DIR / output_file
    
    # Prepare data for Excel
    excel_data = []
    
    for record in data:
        started_at = record.get('started_at')
        time_period = get_time_period(started_at)
        
        # Extract sentiment
        sentiment_text = record.get('sentiment', '')
        if isinstance(sentiment_text, str):
            try:
                sentiment_parsed = json.loads(sentiment_text)
                if isinstance(sentiment_parsed, dict):
                    sentiment_text = sentiment_parsed.get('sentiment_description', sentiment_text)
            except:
                pass  # Use as plain text
        
        # Extract disposition info
        disposition = record.get('disposition', '')
        recommended_action = record.get('recommended_action', '')
        
        # Extract date, start time, end time, and calculate duration
        date_str = ''
        started_time_str = ''
        ended_time_str = ''
        duration_str = ''
        
        ended_at = record.get('ended_at')
        
        # Parse start time and extract date
        if started_at:
            if isinstance(started_at, str):
                try:
                    dt = datetime.strptime(started_at.split('.')[0], '%Y-%m-%d %H:%M:%S')
                    date_str = dt.strftime('%Y-%m-%d')
                    started_time_str = dt.strftime('%H:%M:%S')
                except:
                    date_str = ''
                    started_time_str = ''
            elif isinstance(started_at, datetime):
                date_str = started_at.strftime('%Y-%m-%d')
                started_time_str = started_at.strftime('%H:%M:%S')
        
        # Parse end time
        if ended_at:
            if isinstance(ended_at, str):
                try:
                    dt = datetime.strptime(ended_at.split('.')[0], '%Y-%m-%d %H:%M:%S')
                    ended_time_str = dt.strftime('%H:%M:%S')
                except:
                    ended_time_str = ''
            elif isinstance(ended_at, datetime):
                ended_time_str = ended_at.strftime('%H:%M:%S')
        
        # Calculate duration
        if started_at and ended_at:
            try:
                if isinstance(started_at, str):
                    start_dt = datetime.strptime(started_at.split('.')[0], '%Y-%m-%d %H:%M:%S')
                else:
                    start_dt = started_at
                
                if isinstance(ended_at, str):
                    end_dt = datetime.strptime(ended_at.split('.')[0], '%Y-%m-%d %H:%M:%S')
                else:
                    end_dt = ended_at
                
                duration_seconds = int((end_dt - start_dt).total_seconds())
                minutes = duration_seconds // 60
                seconds = duration_seconds % 60
                duration_str = f"{minutes}m {seconds}s"
            except:
                duration_str = ''
        
        # Build Excel row
        row = {
            'Call Log ID': str(record.get('call_log_id', '')) if record.get('call_log_id') else '',
            'Name': record.get('name', '') or '',
            'Lead Number': record.get('lead_number', '') or '',
            'Date': date_str if date_str else 'N/A',
            'Start Time': started_time_str if started_time_str else 'N/A',
            'End Time': ended_time_str if ended_time_str else 'N/A',
            'Duration': duration_str if duration_str else 'N/A',
            'Time Period': time_period,
            'Summary': record.get('summary', ''),
            'Sentiment': sentiment_text,
            'Key Phrases': format_json_field(record.get('key_phrases')),
            'Key Discussion Points': format_json_field(record.get('key_discussion_points')),
            'Prospect Questions': format_json_field(record.get('prospect_questions')),
            'Prospect Concerns': format_json_field(record.get('prospect_concerns')),
            'Recommendations': record.get('recommendations', ''),
            'Lead Category': record.get('lead_category', ''),
            'Engagement Level': record.get('engagement_level', ''),
            'Disposition': disposition,
            'Recommended Action': recommended_action
        }
        
        excel_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Remove any duplicate columns
    df = df.loc[:, ~df.columns.duplicated()]
    
    # Define column order
    column_order = [
        'Call Log ID',
        'Name',
        'Lead Number',
        'Date',
        'Start Time',
        'End Time',
        'Duration',
        'Time Period',
        'Disposition',
        'Recommended Action',
        'Lead Category',
        'Engagement Level',
        'Sentiment',
        'Summary',
        'Key Phrases',
        'Key Discussion Points',
        'Prospect Questions',
        'Prospect Concerns',
        'Recommendations'
    ]
    
    # Reorder columns (only include columns that exist)
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    # Group by time period for multiple calls
    if len(data) > 1:
        # Multiple sheets - one per time period
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            time_periods = df['Time Period'].unique()
            
            for period in sorted(time_periods):
                period_df = df[df['Time Period'] == period].copy()
                # Remove Time Period column from individual sheets
                period_df = period_df.drop(columns=['Time Period'])
                
                # Create sheet name (Excel sheet names have limitations)
                sheet_name = period[:31].replace(':', '-')
                
                period_df.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_worksheet(writer.sheets[sheet_name], period_df)
    else:
        # Single sheet for single call
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Call Analysis', index=False)
            _format_worksheet(writer.sheets['Call Analysis'], df)
    
    print(f"Excel file created: {output_path}")
    print(f"   Total records exported: {len(excel_data)}")
    
    return str(output_path)


def send_email_with_attachment(
    to_email: str,
    subject: str,
    body: str,
    attachment_path: Optional[str] = None,
    smtp_server: str = None,
    smtp_port: int = None,
    smtp_user: str = None,
    smtp_password: str = None,
    from_email: str = None
) -> bool:
    """Send email with optional Excel file attachment"""
    # Get email settings from environment or use defaults
    smtp_server = smtp_server or os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = smtp_port or int(os.getenv("SMTP_PORT", "587"))
    smtp_user = smtp_user or os.getenv("SMTP_USER")
    smtp_password = smtp_password or os.getenv("SMTP_PASSWORD")
    from_email = from_email or smtp_user or os.getenv("FROM_EMAIL")
    
    if not smtp_user or not smtp_password:
        print("Email credentials not configured. Set SMTP_USER and SMTP_PASSWORD in .env file")
        return False
    
    if attachment_path and not os.path.exists(attachment_path):
        print(f"Attachment file not found: {attachment_path}")
        return False
    
    # Parse recipients - support string (single or comma-separated) or list
    if isinstance(to_email, str):
        recipients = [e.strip() for e in to_email.split(',') if e.strip() and '@' in e.strip()]
    elif isinstance(to_email, list):
        # Handle list - split any items that contain commas
        recipients = []
        for item in to_email:
            if isinstance(item, str):
                # If item contains comma, split it
                if ',' in item:
                    recipients.extend([e.strip() for e in item.split(',') if e.strip() and '@' in e.strip()])
                else:
                    if item.strip() and '@' in item.strip():
                        recipients.append(item.strip())
    else:
        recipients = []
    
    # Remove duplicates while preserving order
    seen = set()
    unique_recipients = []
    for recipient in recipients:
        if recipient not in seen:
            seen.add(recipient)
            unique_recipients.append(recipient)
    recipients = unique_recipients
    
    if not recipients:
        print("No email recipients specified")
        return False
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Add attachment if provided
        if attachment_path:
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        
        # Connect to SMTP server and send
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        
        # Send to all recipients
        failed_recipients = server.sendmail(from_email, recipients, msg.as_string())
        
        server.quit()
        
        if failed_recipients:
            for recipient, error in failed_recipients.items():
                print(f"Failed to send email to {recipient}: {error}")
            return False
        else:
            for recipient in recipients:
                print(f"Email sent successfully to {recipient}")
            return True
        
    except Exception as e:
        print(f"Failed to send email: {str(e)}")
        return False


def process_single_call(call_log_id, recipient_email: Optional[str] = None) -> Dict:
    """
    Process single call: fetch data, create Excel, and send email
    
    Args:
        call_log_id: Call log ID (UUID string or integer row number)
        recipient_email: Optional email address (default: from REPORT_EMAIL env var)
    
    Returns:
        Dictionary with status and file path
    """
    print(f"\n{'='*60}")
    print(f"Processing Single Call Report")
    print(f"{'='*60}")
    print(f"Call Log ID: {call_log_id}\n")
    
    # Fetch call data
    print("Fetching call data from database...")
    call_data = fetch_single_call_data(call_log_id)
    
    if not call_data:
        error_msg = f"Call {call_log_id} not found or has no analysis data"
        print(f"ERROR: {error_msg}")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False
        }
    
    print(f"G?? Found call data")
    
    # Export to Excel
    print("Creating Excel report...")
    try:
        excel_file = export_to_excel([call_data])
        print(f"G?? Excel report created: {excel_file}")
    except Exception as e:
        error_msg = f"Failed to create Excel report: {str(e)}"
        print(f"ERROR: {error_msg}")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False
        }
    
    # Send email
    email_sent = False
    if recipient_email:
        recipients_list = [recipient_email] if isinstance(recipient_email, str) else recipient_email
    else:
        recipients_list = get_email_recipients()
    
    if recipients_list:
        email_subject = f"Single Call Analysis Report - Call ID {call_log_id}"
        email_body = f"""
Hello,

Please find attached the call analysis report for a single call.

Report Details:
- Call Log ID: {call_log_id}
- Total Records: 1
- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Best regards,
Automated Report System
        """
        
        print(f"\nSending email to {len(recipients_list)} recipient(s)...")
        email_sent = send_email_with_attachment(
            to_email=recipients_list,
            subject=email_subject,
            body=email_body,
            attachment_path=excel_file
        )
        
        if email_sent:
            print("G?? Email sent successfully!")
        else:
            print("G?? Email sending failed!")
    else:
        print("No email recipient specified. Set REPORT_EMAIL in .env file or pass recipient_email parameter")
    
    return {
        "status": "success",
        "message": f"Single call report processed successfully",
        "file_path": excel_file,
        "email_sent": email_sent,
        "call_count": 1
    }


def list_all_batch_ids() -> None:
    """List all batch_ids with call counts - helpful for debugging"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = f"""
            SELECT 
                b.id as batch_id,
                b.total_calls,
                COUNT(a.call_log_id) as calls_with_analysis,
                b.created_at as first_call,
                b.updated_at as last_call
            FROM {BATCHES_FULL} b
            LEFT JOIN {BATCH_ENTRIES_FULL} be ON be.batch_id = b.id
            LEFT JOIN {ANALYSIS_FULL} a ON be.call_log_id = a.call_log_id
            GROUP BY b.id, b.total_calls, b.created_at, b.updated_at
            ORDER BY b.created_at DESC
            LIMIT 50
        """
        
        cursor.execute(query)
        batches = cursor.fetchall()
        
        if not batches:
            print("No batch_ids found in database")
            return
        
        print(f"\n{'='*100}")
        print(f"AVAILABLE BATCH IDs (Last 50)")
        print(f"{'='*100}")
        print(f"{'Batch ID':<40} {'Total Calls':<15} {'With Analysis':<15} {'First Call':<20} {'Last Call':<20}")
        print("-" * 100)
        
        for batch_id, total_calls, calls_with_analysis, first_call, last_call in batches:
            batch_str = str(batch_id)
            total_str = str(total_calls)
            analysis_str = f"{calls_with_analysis}/{total_calls}"
            first_str = first_call.strftime('%Y-%m-%d %H:%M:%S') if first_call else 'N/A'
            last_str = last_call.strftime('%Y-%m-%d %H:%M:%S') if last_call else 'N/A'
            
            print(f"{batch_str:<40} {total_str:<15} {analysis_str:<15} {first_str:<20} {last_str:<20}")
        
        print("-" * 100)
        print(f"\nTotal batches shown: {len(batches)}")
        print(f"\nTo export a batch, use:")
        print(f"  python call_report_service.py --batch-id <batch_id>")
        
    finally:
        cursor.close()
        conn.close()


def process_bulk_calls(batch_id: str, recipient_email: Optional[str] = None) -> Dict:
    """
    Process bulk calls by batch_id: fetch all calls, create Excel, and send email
    
    Args:
        batch_id: Batch ID (UUID) from call_logs_voiceagent.batch_id column
        recipient_email: Optional email address (default: from REPORT_EMAIL env var)
    
    Returns:
        Dictionary with status and file path
    """
    print(f"\n{'='*60}")
    print(f"Processing Bulk Call Report")
    print(f"{'='*60}")
    print(f"Batch ID: {batch_id}\n")
    
    # First check if batch_id exists and get diagnostics
    print("Checking batch_id in database...")
    batch_info = check_batch_id_exists(batch_id)
    
    if not batch_info["exists"]:
        error_msg = f"Batch ID {batch_id} not found in database"
        print(f"ERROR: {error_msg}")
        print(f"   No calls found with this batch_id.")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False,
            "call_count": 0
        }
    
    print(f"G?? Batch ID exists: {batch_info['total_calls']} total call(s)")
    print(f"  - Calls with analysis: {batch_info['calls_with_analysis']}")
    print(f"  - Calls without analysis: {batch_info['calls_without_analysis']}")
    
    # Check if any calls have analysis
    if batch_info["calls_with_analysis"] == 0:
        error_msg = f"Batch ID {batch_id} has {batch_info['total_calls']} call(s) but none have analysis data yet"
        print(f"\nERROR: {error_msg}")
        print(f"   Analysis may still be in progress. Please wait for analysis to complete.")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False,
            "call_count": batch_info["total_calls"],
            "calls_without_analysis": batch_info["calls_without_analysis"]
        }
    
    # Fetch all calls for this batch
    print(f"\nFetching calls with analysis data...")
    calls_data = fetch_bulk_call_data_by_batch_id(batch_id)
    
    if not calls_data:
        error_msg = f"Failed to fetch analysis data for batch_id {batch_id}"
        print(f"ERROR: {error_msg}")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False,
            "call_count": 0
        }
    
    print(f"G?? Found {len(calls_data)} call(s) with analysis data")
    
    # Warn if some calls are missing analysis
    if batch_info["calls_without_analysis"] > 0:
        print(f"G?? Warning: {batch_info['calls_without_analysis']} call(s) in this batch don't have analysis yet (will be skipped)")
    
    # Export to Excel
    print("Creating Excel report...")
    try:
        excel_file = export_to_excel(calls_data, batch_id=batch_id)
        print(f"G?? Excel report created: {excel_file}")
    except Exception as e:
        error_msg = f"Failed to create Excel report: {str(e)}"
        print(f"ERROR: {error_msg}")
        return {
            "status": "error",
            "message": error_msg,
            "file_path": None,
            "email_sent": False,
            "call_count": len(calls_data)
        }
    
    # Send email
    email_sent = False
    if recipient_email:
        recipients_list = [recipient_email] if isinstance(recipient_email, str) else recipient_email
    else:
        recipients_list = get_email_recipients()
    
    if recipients_list:
        email_subject = f"Bulk Call Analysis Report - Batch ID {batch_id} ({len(calls_data)} Calls)"
        email_body = f"""
Hello,

Please find attached the call analysis report for bulk calls.

Report Details:
- Batch ID: {batch_id}
- Total Records: {len(calls_data)}
- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

The report is organized by time periods in separate Excel sheets.

Best regards,
Automated Report System
        """
        
        print(f"\nSending email to {len(recipients_list)} recipient(s)...")
        email_sent = send_email_with_attachment(
            to_email=recipients_list,
            subject=email_subject,
            body=email_body,
            attachment_path=excel_file
        )
        
        if email_sent:
            print("G?? Email sent successfully!")
        else:
            print("G?? Email sending failed!")
    else:
        print("No email recipient specified. Set REPORT_EMAIL in .env file or pass recipient_email parameter")
    
    return {
        "status": "success",
        "message": f"Bulk call report processed successfully",
        "file_path": excel_file,
        "email_sent": email_sent,
        "call_count": len(calls_data),
        "batch_id": batch_id
    }


def main():
    """Main CLI entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Call Report Service - Single and Bulk Call Reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # List all available batch_ids
    python call_report_service.py --list-batches
    
    # Single call report
    python call_report_service.py --call-id 123
    
    # Single call report with email
    python call_report_service.py --call-id 123 --email recipient@example.com
    
    # Bulk call report by batch_id
    python call_report_service.py --batch-id 6241c2f0-b637-4fe5-8e5c-95dd2faf7db4
    
    # Bulk call report with email
    python call_report_service.py --batch-id 6241c2f0-b637-4fe5-8e5c-95dd2faf7db4 --email recipient@example.com
        """
    )
    
    parser.add_argument(
        '--list-batches',
        action='store_true',
        help='List all available batch_ids with call counts'
    )
    
    input_group = parser.add_mutually_exclusive_group(required=False)
    input_group.add_argument(
        '--call-id',
        help='Call log ID for single call report (UUID string or integer row number)'
    )
    input_group.add_argument(
        '--batch-id',
        help='Batch ID (UUID) for bulk call report from call_logs_voiceagent.batch_id'
    )
    
    parser.add_argument(
        '--email',
        help='Email address to send report to (optional, defaults to REPORT_EMAIL from .env)',
        default=None
    )
    
    args = parser.parse_args()
    
    # Handle --list-batches option
    if args.list_batches:
        list_all_batch_ids()
        return 0
    
    # Require either --call-id or --batch-id if --list-batches is not used
    if not args.call_id and not args.batch_id:
        parser.error("Either --call-id, --batch-id, or --list-batches is required")
    
    # Get recipient email
    recipient_email = args.email or (','.join(get_email_recipients()) if get_email_recipients() else None)
    
    try:
        if args.call_id:
            # Process single call
            result = process_single_call(args.call_id, recipient_email)
            
            print(f"\n{'='*60}")
            print("REPORT SUMMARY")
            print(f"{'='*60}")
            print(f"Status: {result['status']}")
            print(f"Message: {result['message']}")
            print(f"File Path: {result.get('file_path', 'N/A')}")
            print(f"Email Sent: {'Yes' if result.get('email_sent') else 'No'}")
            print(f"Call Count: {result.get('call_count', 0)}")
            print(f"{'='*60}\n")
            
        elif args.batch_id:
            # Process bulk calls
            result = process_bulk_calls(args.batch_id, recipient_email)
            
            print(f"\n{'='*60}")
            print("REPORT SUMMARY")
            print(f"{'='*60}")
            print(f"Status: {result['status']}")
            print(f"Message: {result['message']}")
            print(f"File Path: {result.get('file_path', 'N/A')}")
            print(f"Email Sent: {'Yes' if result.get('email_sent') else 'No'}")
            print(f"Call Count: {result.get('call_count', 0)}")
            print(f"Batch ID: {result.get('batch_id', 'N/A')}")
            print(f"{'='*60}\n")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())


