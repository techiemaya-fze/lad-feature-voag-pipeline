"""
Batch Report Service - Excel Report Generation for Completed Batches

This service generates Excel reports for completed batch calls and can optionally
send them via email. It integrates with the existing database infrastructure
and is called from main.py when a batch completes.

Supports two email methods:
- SMTP (default): Traditional SMTP email sending
- OAuth: Uses Google Gmail API with stored OAuth tokens (bypasses SMTP blocks)

Usage from main.py:
    from analysis.batch_report_service import generate_batch_report
    
    # When batch completes (completed, stopped, or cancelled)
    result = await generate_batch_report(batch_id, send_email=True)

CLI Usage:
    # List all batches
    uv run python -m post_call_analysis.batch_report_service --list-batches
    
    # Generate report for a batch
    uv run python -m post_call_analysis.batch_report_service --batch-id <UUID>
    
    # Generate and email report
    uv run python -m post_call_analysis.batch_report_service --batch-id <UUID> --send-email

Environment Variables Required:
    - DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD (database)
    - BATCH_REPORT_EMAIL or REPORT_EMAIL (optional, for email reports)
      Special value 'self' can be included in the list to automatically include
      the connected Gmail address of the OAuth user (BATCH_EMAIL_OAUTH_USER_ID)
      Example: REPORT_EMAIL=self,user1@example.com,user2@example.com
    
    For SMTP (default):
    - SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD
    
    For OAuth (set BATCH_EMAIL_METHOD=oauth):
    - BATCH_EMAIL_METHOD=oauth
    - BATCH_EMAIL_OAUTH_USER_ID=10 (user ID with Google OAuth tokens)
"""

import os
import json
import asyncio
import logging
import base64
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
from pathlib import Path

from dotenv import load_dotenv

# Gulf Standard Time (GST = UTC+4)
GST = timezone(timedelta(hours=4))

load_dotenv()

# Schema constants for table names
from db.schema_constants import (
    CALL_LOGS_FULL,
    ANALYSIS_FULL,
    LEADS_FULL,
    BATCHES_FULL,
    BATCH_ENTRIES_FULL,
)

# Configure logging
logger = logging.getLogger("post_call_analysis.batch_report_service")

# Optional imports for Excel and email
try:
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("pandas/openpyxl not installed. Excel export disabled. Install with: pip install pandas openpyxl")

try:
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    EMAIL_AVAILABLE = True
except ImportError:
    EMAIL_AVAILABLE = False

# OAuth email support (Gmail API)
try:
    from utils.google_credentials import GoogleCredentialResolver, GoogleCredentialError
    from tools.gmail_email_tool import GmailEmailTool, EmailPayload, GmailToolError
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    OAUTH_EMAIL_AVAILABLE = True
except ImportError:
    OAUTH_EMAIL_AVAILABLE = False
    logger.warning("OAuth email not available (missing google_credentials or gmail_email_tool)")

# Use existing connection pool infrastructure
try:
    from db.connection_pool import get_raw_connection, return_connection, USE_CONNECTION_POOLING
    DB_POOL_AVAILABLE = True
except ImportError:
    DB_POOL_AVAILABLE = False
    logger.warning("Connection pool not available, using direct connections")
    try:
        import psycopg2
    except ImportError:
        logger.error("psycopg2 not installed")

# Create exports directory
EXPORTS_DIR = Path(__file__).parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


def _convert_to_gst(dt: Optional[datetime]) -> Optional[datetime]:
    """
    Convert datetime to Gulf Standard Time (UTC+4)
    
    Args:
        dt: datetime object (aware or naive, assumes UTC if naive)
    
    Returns:
        datetime in GST timezone
    """
    if dt is None:
        return None
    
    # If datetime is naive (no timezone), assume it's UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    
    # Convert to GST (UTC+4)
    return dt.astimezone(GST)


def _get_db_config() -> Dict[str, Any]:
    """Get database configuration from environment"""
    return {
        "host": os.getenv("DB_HOST", "localhost"),
        "port": int(os.getenv("DB_PORT", "5432")),
        "database": os.getenv("DB_NAME", "salesmaya_agent"),
        "user": os.getenv("DB_USER", "postgres"),
        "password": os.getenv("DB_PASSWORD"),
        "connect_timeout": 30,
    }


def _get_connection():
    """Get database connection using pool if available (raw connection, must be returned manually)"""
    db_config = _get_db_config()
    if DB_POOL_AVAILABLE:
        return get_raw_connection(db_config)
    else:
        import psycopg2
        return psycopg2.connect(**db_config)


def _return_conn(conn):
    """Return connection to pool if pooling is enabled"""
    if DB_POOL_AVAILABLE and USE_CONNECTION_POOLING:
        return_connection(conn, _get_db_config())
    else:
        conn.close()


def get_email_recipients(env_key: str = "BATCH_REPORT_EMAIL") -> List[str]:
    """
    Get list of email recipients from environment variable.
    
    Supports a special 'self' keyword which will be resolved to the connected
    Gmail address of the OAuth user (BATCH_EMAIL_OAUTH_USER_ID) when 
    get_email_recipients_async() is called.
    
    Args:
        env_key: Environment variable name (default: BATCH_REPORT_EMAIL)
    
    Returns:
        List of email addresses (may include 'self' placeholder)
    """
    emails_str = os.getenv(env_key, "").strip()
    if not emails_str:
        # Fall back to REPORT_EMAIL for backwards compatibility
        emails_str = os.getenv("REPORT_EMAIL", "").strip()
    
    if not emails_str:
        return []
    
    # Remove quotes if present
    emails_str = emails_str.strip('"\'')
    
    # Split by comma, semicolon, or newline
    for sep in [',', ';', '\n']:
        if sep in emails_str:
            emails = [e.strip() for e in emails_str.split(sep)]
            break
    else:
        emails = [emails_str.strip()]
    
    # Filter: keep valid emails OR the special 'self' keyword
    return [e for e in emails if e and ('@' in e or e.lower() == 'self')]


async def _get_oauth_user_connected_gmail(user_id: Optional[int] = None) -> Optional[str]:
    """
    Get the connected Gmail address for the specified user or the default OAuth user.
    
    Args:
        user_id: Optional user ID to look up. Falls back to BATCH_EMAIL_OAUTH_USER_ID env var.
    
    Returns:
        Connected Gmail address or None if not found/configured
    """
    # Use provided user_id or fall back to env var
    if user_id is not None:
        oauth_user_id = user_id
        logger.info(f"Resolving 'self' email using provided user_id={oauth_user_id} (batch initiator)")
    else:
        oauth_user_id_str = os.getenv("BATCH_EMAIL_OAUTH_USER_ID", "")
        if not oauth_user_id_str:
            logger.warning("No user_id provided and BATCH_EMAIL_OAUTH_USER_ID not set, cannot resolve 'self' email recipient")
            return None
        
        try:
            oauth_user_id = int(oauth_user_id_str)
            logger.info(f"Resolving 'self' email using fallback BATCH_EMAIL_OAUTH_USER_ID={oauth_user_id}")
        except ValueError:
            logger.error(f"Invalid BATCH_EMAIL_OAUTH_USER_ID: {oauth_user_id_str}")
            return None
    
    # Query the database for the connected_gmail
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT email FROM lad_dev.users WHERE id = %s",
                (oauth_user_id,)
            )
            row = cur.fetchone()
            if row and row[0]:
                connected_gmail = row[0]
                logger.info(f"Resolved 'self' to connected_gmail: {connected_gmail} (user_id={oauth_user_id})")
                return connected_gmail
            else:
                logger.warning(f"No connected_gmail found for user_id={oauth_user_id}")
                return None
    except Exception as e:
        logger.error(f"Failed to fetch connected_gmail for user {oauth_user_id}: {e}")
        return None
    finally:
        _return_conn(conn)


async def get_email_recipients_async(
    env_key: str = "BATCH_REPORT_EMAIL",
    initiated_by: Optional[int] = None,
) -> List[str]:
    """
    Get list of email recipients from environment variable, resolving 'self' to 
    the connected Gmail of the batch initiator or OAuth user.
    
    If 'self' is present in the recipient list, it will be replaced with the 
    connected_gmail address of:
    1. The initiated_by user (if provided) - the person who started the batch
    2. The user specified by BATCH_EMAIL_OAUTH_USER_ID (fallback)
    
    Args:
        env_key: Environment variable name (default: BATCH_REPORT_EMAIL)
        initiated_by: User ID who initiated the batch (used to resolve 'self')
    
    Returns:
        List of resolved email addresses
    """
    raw_recipients = get_email_recipients(env_key)
    
    # Check if 'self' is in the list
    has_self = any(e.lower() == 'self' for e in raw_recipients)
    
    if not has_self:
        return raw_recipients
    
    # Resolve 'self' to the initiator's connected Gmail (or fallback to env var user)
    connected_gmail = await _get_oauth_user_connected_gmail(user_id=initiated_by)
    
    resolved_recipients = []
    for email in raw_recipients:
        if email.lower() == 'self':
            if connected_gmail:
                resolved_recipients.append(connected_gmail)
            else:
                logger.warning("'self' in REPORT_EMAIL but could not resolve connected_gmail - skipping")
        else:
            resolved_recipients.append(email)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_recipients = []
    for email in resolved_recipients:
        email_lower = email.lower()
        if email_lower not in seen:
            seen.add(email_lower)
            unique_recipients.append(email)
    
    return unique_recipients


async def get_batch_info(batch_id: str) -> Optional[Dict]:
    """
    Get batch information from database
    
    Args:
        batch_id: Batch UUID
    
    Returns:
        Batch info dict or None if not found
    """
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            # v2 schema: lad_dev.voice_call_batches uses finished_at, no cancelled_calls
            cur.execute(f"""
                SELECT 
                    id, metadata->>'job_id' as job_id, status, total_calls, completed_calls,
                    failed_calls, created_at, finished_at,
                    agent_id, initiated_by_user_id
                FROM {BATCHES_FULL}
                WHERE id = %s::uuid
            """, (str(batch_id),))
            
            row = cur.fetchone()
            if not row:
                return None
            
            return {
                "id": str(row[0]),
                "job_id": row[1],
                "status": row[2],
                "total_calls": row[3],
                "completed_calls": row[4],
                "failed_calls": row[5],
                "cancelled_calls": 0,  # v2 schema doesn't have this, default to 0
                "created_at": row[6],
                "completed_at": row[7],  # Using finished_at column
                "agent_id": row[8],
                "initiated_by": row[9],
            }
    finally:
        _return_conn(conn)


async def fetch_batch_call_data(batch_id: str) -> List[Dict]:
    """
    Fetch all call analysis data for a batch
    
    Args:
        batch_id: Batch UUID
    
    Returns:
        List of call data dictionaries
    """
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            # Check if the leads table actually has a 'metadata' column to avoid SQL errors
            schema_name, table_name = None, None
            if '.' in LEADS_FULL:
                schema_name, table_name = LEADS_FULL.split('.', 1)

            has_metadata_col = False
            if schema_name and table_name:
                cur.execute(
                    "SELECT EXISTS(SELECT 1 FROM information_schema.columns WHERE table_schema = %s AND table_name = %s AND column_name = %s)",
                    (schema_name, table_name, 'metadata')
                )
                row_exists = cur.fetchone()
                has_metadata_col = bool(row_exists[0]) if row_exists is not None else False

            # Build safe lead_name expression depending on whether metadata JSON column exists
            # Use first_name + last_name from leads table for lead_name (no metadata lookup)
            # Handle NULL values properly: COALESCE each field before concatenation to avoid NULL result
            lead_name_expr = "TRIM(COALESCE(l.first_name, '') || ' ' || COALESCE(l.last_name, ''))"

            sql = f"""
                SELECT DISTINCT ON (cl.id)
                    cl.id as call_log_id,
                    cl.started_at,
                    cl.ended_at,
                    cl.status as call_status,
                    cl.duration_seconds as call_duration,
                    cl.lead_id,
                    {lead_name_expr} as lead_name,
                    COALESCE(l.phone, '') as lead_number,
                    a.summary,
                    a.sentiment,
                    a.key_points,
                    a.disposition,
                    a.recommended_action,
                    a.lead_category,
                    a.engagement_level,
                    a.prospect_questions,
                    a.prospect_concerns,
                    a.recommendations
                FROM {BATCH_ENTRIES_FULL} be
                INNER JOIN {CALL_LOGS_FULL} cl ON be.call_log_id = cl.id
                LEFT JOIN {ANALYSIS_FULL} a ON a.call_log_id = cl.id
                LEFT JOIN {LEADS_FULL} l ON cl.lead_id = l.id
                WHERE be.batch_id = %s::uuid
                ORDER BY cl.id, a.created_at DESC NULLS LAST, cl.started_at ASC
            """

            cur.execute(sql, (str(batch_id),))

            
            columns = [desc[0] for desc in cur.description]
            rows = cur.fetchall()
            
            data = []
            for row in rows:
                row_dict = dict(zip(columns, row))
                
                # Create a clean record to avoid any data corruption
                clean_record = {
                    'call_log_id': row_dict.get('call_log_id'),
                    'started_at': row_dict.get('started_at'),
                    'ended_at': row_dict.get('ended_at'),
                    'call_status': row_dict.get('call_status'),
                    'call_duration': row_dict.get('call_duration'),
                    'lead_id': row_dict.get('lead_id'),
                    'lead_name': row_dict.get('lead_name', ''),
                    'lead_number': row_dict.get('lead_number', ''),
                    'summary': row_dict.get('summary', ''),
                    'sentiment': row_dict.get('sentiment', ''),
                    'disposition': row_dict.get('disposition', ''),
                    'recommended_action': row_dict.get('recommended_action', ''),
                    'lead_category': row_dict.get('lead_category', ''),
                    'engagement_level': row_dict.get('engagement_level', ''),
                    'recommendations': row_dict.get('recommendations', ''),
                }

                # We intentionally do not query JSON metadata here — use leads.first_name/last_name
                
                # Handle key_points (JSONB array)
                key_points = row_dict.get('key_points') or []
                if isinstance(key_points, str):
                    try:
                        key_points = json.loads(key_points)
                    except json.JSONDecodeError:
                        key_points = []
                clean_record['key_discussion_points'] = key_points
                
                # Handle prospect questions (JSONB array)
                prospect_questions = row_dict.get('prospect_questions') or []
                if isinstance(prospect_questions, str):
                    try:
                        prospect_questions = json.loads(prospect_questions)
                    except json.JSONDecodeError:
                        prospect_questions = []
                clean_record['prospect_questions'] = prospect_questions
                
                # Handle prospect concerns (JSONB array)
                prospect_concerns = row_dict.get('prospect_concerns') or []
                if isinstance(prospect_concerns, str):
                    try:
                        prospect_concerns = json.loads(prospect_concerns)
                    except json.JSONDecodeError:
                        prospect_concerns = []
                clean_record['prospect_concerns'] = prospect_concerns
                
                # Handle sentiment
                clean_record['sentiment_description'] = clean_record['sentiment']
                
                data.append(clean_record)
            
            return data
    finally:
        _return_conn(conn)


async def get_batch_call_count(batch_id: str) -> Dict[str, int]:
    """
    Get count of total calls and analyzed calls for a batch
    
    Args:
        batch_id: Batch UUID
    
    Returns:
        Dict with 'total_calls' and 'analyzed_calls' counts
    """
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT 
                    COUNT(be.id) as total_calls,
                    COUNT(a.id) as analyzed_calls
                FROM {BATCH_ENTRIES_FULL} be
                LEFT JOIN {ANALYSIS_FULL} a ON a.call_log_id = be.call_log_id
                WHERE be.batch_id = %s::uuid
            """, (str(batch_id),))
            
            row = cur.fetchone()
            return {
                "total_calls": row[0] if row else 0,
                "analyzed_calls": row[1] if row else 0,
            }
    finally:
        _return_conn(conn)


async def wait_for_analysis_completion(
    batch_id: str,
    max_wait_seconds: int = 120,
    poll_interval_seconds: int = 10
) -> Dict[str, int]:
    """
    Wait for post-call analysis to complete for all calls in a batch
    
    Args:
        batch_id: Batch UUID
        max_wait_seconds: Maximum time to wait (default 2 minutes)
        poll_interval_seconds: How often to check (default 10 seconds)
    
    Returns:
        Dict with 'total_calls' and 'analyzed_calls' counts
    """
    logger.info(f"Waiting for analysis completion for batch {batch_id} (max {max_wait_seconds}s)")
    
    start_time = datetime.now(timezone.utc)
    last_counts = None
    stable_count = 0  # Track how many times counts have been stable
    
    while True:
        counts = await get_batch_call_count(batch_id)
        elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
        
        logger.info(f"Analysis status: {counts['analyzed_calls']}/{counts['total_calls']} analyzed ({elapsed:.0f}s elapsed)")
        
        # All calls analyzed
        if counts['total_calls'] > 0 and counts['analyzed_calls'] >= counts['total_calls']:
            logger.info(f"All {counts['total_calls']} calls have been analyzed")
            return counts
        
        # Check for stable counts (analysis might be stuck or some calls don't get analyzed)
        if last_counts and counts['analyzed_calls'] == last_counts['analyzed_calls']:
            stable_count += 1
            # If counts have been stable for 3 polls, assume analysis is complete
            if stable_count >= 3:
                logger.warning(f"Analysis counts stable for {stable_count * poll_interval_seconds}s - proceeding with {counts['analyzed_calls']}/{counts['total_calls']} analyzed")
                return counts
        else:
            stable_count = 0
        
        last_counts = counts
        
        # Timeout
        if elapsed >= max_wait_seconds:
            logger.warning(f"Timeout waiting for analysis: {counts['analyzed_calls']}/{counts['total_calls']} analyzed after {elapsed:.0f}s")
            return counts
        
        await asyncio.sleep(poll_interval_seconds)


async def list_batches(limit: int = 50) -> List[Dict]:
    """
    List recent batches with summary info
    
    Args:
        limit: Maximum number of batches to return
    
    Returns:
        List of batch summary dictionaries
    """
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT 
                    id, metadata->>'job_id' as job_id, status, total_calls, completed_calls,
                    failed_calls, created_at, finished_at
                FROM {BATCHES_FULL}
                ORDER BY created_at DESC
                LIMIT %s
            """, (limit,))
            
            batches = []
            for row in cur.fetchall():
                batches.append({
                    "id": str(row[0]),
                    "job_id": row[1],
                    "status": row[2],
                    "total_calls": row[3],
                    "completed_calls": row[4],
                    "failed_calls": row[5],
                    "cancelled_calls": 0,  # Default to 0 since column doesn't exist
                    "created_at": row[6],
                    "completed_at": row[7],
                })
            
            return batches
    finally:
        _return_conn(conn)


def _get_time_period(started_at: datetime) -> str:
    """Categorize call into time period based on start time (in GST)"""
    if not started_at:
        return "Unknown"
    
    # Convert to GST for time period categorization
    gst_time = _convert_to_gst(started_at)
    if not gst_time:
        return "Unknown"
    
    hour = gst_time.hour
    
    if 9 <= hour < 12:
        return "09:00-12:00 (Morning)"
    elif 12 <= hour < 14:
        return "12:00-14:00 (Midday)"
    elif 14 <= hour < 17:
        return "14:00-17:00 (Afternoon)"
    elif 17 <= hour < 20:
        return "17:00-20:00 (Evening)"
    elif 20 <= hour < 24:
        return "20:00-24:00 (Night)"
    else:
        return "00:00-09:00 (Early Morning)"


def _format_list_field(value) -> str:
    """Format list/JSON field for Excel display"""
    if isinstance(value, list):
        if not value:
            return ""
        # Format as bullet points for better readability in Excel
        bullet_points = [f"• {str(item)}" for item in value if str(item).strip()]
        return "\n".join(bullet_points)
    elif isinstance(value, dict):
        return json.dumps(value, indent=2)
    elif value is None:
        return ""
    return str(value)


def _format_duration(started_at, ended_at) -> str:
    """Format duration from timestamps"""
    if not started_at or not ended_at:
        return "N/A"
    
    try:
        if isinstance(started_at, str):
            started_at = datetime.fromisoformat(started_at.replace('Z', '+00:00'))
        if isinstance(ended_at, str):
            ended_at = datetime.fromisoformat(ended_at.replace('Z', '+00:00'))
        
        # Ensure both are timezone-aware for accurate calculation
        if started_at.tzinfo is None:
            started_at = started_at.replace(tzinfo=timezone.utc)
        if ended_at.tzinfo is None:
            ended_at = ended_at.replace(tzinfo=timezone.utc)
        
        duration_seconds = int((ended_at - started_at).total_seconds())
        if duration_seconds < 0:
            return "N/A"
        minutes = duration_seconds // 60
        seconds = duration_seconds % 60
        return f"{minutes}m {seconds}s"
    except Exception:
        return "N/A"


def _apply_excel_formatting(worksheet, df):
    """Apply professional formatting to Excel worksheet"""
    if not EXCEL_AVAILABLE:
        return
    
    # Define professional color scheme
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Professional dark blue
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray alternate rows
    
    # Define enhanced fonts
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    data_font = Font(size=10, name="Calibri")
    
    # Define professional borders
    thick_border = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79')
    )
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Define alignments
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format header row with enhanced styling
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thick_border
    
    worksheet.row_dimensions[1].height = 35  # Taller header for better appearance
    
    # Format data rows with alternating colors and enhanced styling
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        is_even_row = (row_idx - 2) % 2 == 0
        
        for col_idx, cell in enumerate(row, start=1):
            # Apply alternating row colors
            if is_even_row:
                cell.fill = alt_row_fill
            
            cell.font = data_font
            cell.border = thin_border
            
            # Special formatting for specific columns
            col_letter = cell.column_letter
            if col_letter in ['A']:  # Call Log ID - center aligned
                cell.alignment = center_alignment
            elif col_letter in ['D', 'E', 'F', 'G']:  # Date/Time columns - center aligned
                cell.alignment = center_alignment
            elif col_letter in ['H', 'I', 'J', 'K', 'L', 'M']:  # Status and category columns - center aligned
                cell.alignment = center_alignment
            else:  # Text columns - left aligned with wrap
                cell.alignment = data_alignment
        
        # Set row height for better readability
        worksheet.row_dimensions[row_idx].height = 25
    
    # Enhanced column widths for better presentation
    column_widths = {
        'Call Log ID': 25,
        'Lead Name': 25,
        'Lead Number': 18,
        'Date': 15,
        'Start Time (GST)': 18,
        'End Time (GST)': 18,
        'Duration': 12,
        'Call Status': 15,
        'Time Period (GST)': 25,
        'Disposition': 25,
        'Recommended Action': 45,
        'Lead Category': 20,
        'Engagement Level': 18,
        'Sentiment': 30,
        'Summary': 70,
        'Key Discussion Points': 55,
        'Prospect Questions': 55,
        'Prospect Concerns': 55,
        'Recommendations': 55,
    }
    
    # Apply column widths and additional formatting
    for idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        
        # Set custom widths or calculate optimal width
        if col in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col]
        else:
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 3, 70)
    
    # Freeze header row and first column for better navigation
    worksheet.freeze_panes = 'B2'
    
    # Enable advanced filtering
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # Add a title row above headers (optional enhancement)
    worksheet.insert_rows(1)
    title_cell = worksheet['A1']
    title_cell.value = f"Batch Call Analysis Report"
    title_cell.font = Font(bold=True, size=16, color="1F4E79", name="Calibri")
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Merge title across several columns for better appearance
    worksheet.merge_cells('A1:F1')
    worksheet.row_dimensions[1].height = 30
    
    # Update freeze panes to account for title row
    worksheet.freeze_panes = 'B3'


def _apply_enhanced_excel_formatting(worksheet, df):
    """Apply enterprise-grade formatting inspired by the provided HTML template"""
    if not EXCEL_AVAILABLE:
        return
    
    # Professional color palette (matching HTML template design)
    HEADER_DARK = "0F172A"      # Dark slate (matching template gradient start)
    HEADER_GRADIENT = "1E293B"  # Medium slate (matching template gradient middle)
    ACCENT_BLUE = "3B82F6"      # Professional blue (matching template accent)
    BACKGROUND_WHITE = "FFFFFF" # Clean white
    ALT_ROW_LIGHT = "F8FAFC"    # Very light blue-gray (matching template)
    BORDER_LIGHT = "E2E8F0"     # Light border (matching template)
    TEXT_DARK = "0F172A"        # Dark text (matching template)
    TEXT_MUTED = "64748B"       # Muted text (matching template)
    
    # Status badge colors (matching HTML template status badges)
    STATUS_SUCCESS = "D1FAE5"   # Light green background
    STATUS_SUCCESS_TEXT = "065F46"  # Dark green text
    STATUS_WARNING = "FEF3C7"   # Light amber background  
    STATUS_WARNING_TEXT = "92400E"  # Dark amber text
    STATUS_DANGER = "FEE2E2"    # Light red background
    STATUS_DANGER_TEXT = "991B1B"   # Dark red text
    
    # Professional typography (matching HTML template Inter/Segoe UI hierarchy)
    title_font = Font(bold=True, size=20, color="FFFFFF", name="Inter")
    subtitle_font = Font(size=11, color="CBD5E1", name="Inter")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Inter")
    data_font = Font(size=10, color=TEXT_DARK, name="Inter")
    
    # Professional fill patterns
    header_fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    alt_row_fill = PatternFill(start_color=ALT_ROW_LIGHT, end_color=ALT_ROW_LIGHT, fill_type="solid")
    
    # Clean border styles
    header_border = Border(
        left=Side(style='thin', color=HEADER_DARK),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='thin', color=HEADER_DARK),
        bottom=Side(style='thin', color=HEADER_DARK)
    )
    data_border = Border(
        left=Side(style='thin', color=BORDER_LIGHT),
        right=Side(style='thin', color=BORDER_LIGHT),
        top=Side(style='hair', color=BORDER_LIGHT),
        bottom=Side(style='thin', color=BORDER_LIGHT)
    )
    
    # Professional alignments
    title_alignment = Alignment(horizontal='left', vertical='center')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Create executive dashboard header (matching HTML template)
    worksheet.insert_rows(1, 5)  # Insert 5 rows for enhanced header section
    
    # Company header (matching template's "APEX CORPORATION" style)
    company_cell = worksheet['A1']
    company_cell.value = "VOICE AGENT ANALYTICS PLATFORM"
    company_cell.font = Font(bold=True, size=10, color=ACCENT_BLUE, name="Inter")
    company_cell.alignment = title_alignment
    company_cell.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    worksheet.merge_cells('A1:T1')
    worksheet.row_dimensions[1].height = 20
    
    # Main title (matching template's main heading style)
    title_cell = worksheet['A2']
    title_cell.value = "Voice Agent Performance Dashboard Report"
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    title_cell.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    worksheet.merge_cells('A2:T2')
    worksheet.row_dimensions[2].height = 40
    
    # Subtitle (matching template's executive summary style)
    subtitle_cell = worksheet['A3']
    subtitle_cell.value = f"Executive Summary of Voice Agent Performance & Lead Conversion Analytics • Generated {datetime.now(GST).strftime('%B %d, %Y at %H:%M GST')}"
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = title_alignment
    subtitle_cell.fill = PatternFill(start_color=HEADER_DARK, end_color=HEADER_DARK, fill_type="solid")
    worksheet.merge_cells('A3:T3')
    worksheet.row_dimensions[3].height = 28
    
    # Separator rows for clean spacing
    worksheet.row_dimensions[4].height = 12
    worksheet.row_dimensions[5].height = 8
    
    # Format header row (now row 6) with executive table styling
    header_row = 6
    for col_idx, cell in enumerate(worksheet[header_row], start=1):
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Light gray header
        cell.font = Font(bold=True, size=11, color="475569", name="Inter")  # Template header text color
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin', color=BORDER_LIGHT),
            right=Side(style='thin', color=BORDER_LIGHT),
            top=Side(style='medium', color=BORDER_LIGHT),
            bottom=Side(style='medium', color=BORDER_LIGHT)
        )
    
    worksheet.row_dimensions[header_row].height = 45
    
    # Format data rows with executive table styling (matching HTML template hover effects)
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row+1, max_row=worksheet.max_row), start=header_row+1):
        is_even_row = (row_idx - header_row - 1) % 2 == 0
        
        for col_idx, cell in enumerate(row, start=1):
            # Apply clean alternating row design (matching HTML template)
            if not is_even_row:  # Odd rows get light background
                cell.fill = PatternFill(start_color=ALT_ROW_LIGHT, end_color=ALT_ROW_LIGHT, fill_type="solid")
            
            cell.font = Font(size=10, color="334155", name="Inter")  # Template body text color
            cell.border = Border(
                left=Side(style='hair', color=BORDER_LIGHT),
                right=Side(style='hair', color=BORDER_LIGHT),
                top=Side(style='hair', color="F1F5F9"),
                bottom=Side(style='thin', color="F1F5F9")
            )
            
            # Column-specific formatting (matching HTML template badge system)
            if col_idx <= len(df.columns):
                col_name = df.columns[col_idx-1]
                
                # Status badge formatting (exactly like HTML template)
                if col_name == 'Call Status':
                    if str(cell.value).lower() == 'completed':
                        cell.fill = PatternFill(start_color=STATUS_SUCCESS, end_color=STATUS_SUCCESS, fill_type="solid")
                        cell.font = Font(color=STATUS_SUCCESS_TEXT, bold=True, size=10, name="Inter")
                    elif str(cell.value).lower() in ['failed', 'error']:
                        cell.fill = PatternFill(start_color=STATUS_DANGER, end_color=STATUS_DANGER, fill_type="solid")
                        cell.font = Font(color=STATUS_DANGER_TEXT, bold=True, size=10, name="Inter")
                    elif str(cell.value).lower() in ['in-progress', 'pending']:
                        cell.fill = PatternFill(start_color=STATUS_WARNING, end_color=STATUS_WARNING, fill_type="solid")
                        cell.font = Font(color=STATUS_WARNING_TEXT, bold=True, size=10, name="Inter")
                    cell.alignment = center_alignment
                    
                # Clean formatting for different data types
                elif col_name in ['Call Log ID']:
                    cell.alignment = center_alignment
                    cell.font = Font(size=9, color=TEXT_MUTED, name="Courier New")  # Monospace like HTML template
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Light background
                    
                elif col_name in ['Date', 'Start Time (GST)', 'End Time (GST)', 'Duration', 'Time Period (GST)']:
                    cell.alignment = center_alignment
                    cell.font = Font(size=10, name="Segoe UI")
                    
                elif col_name in ['Lead Name', 'Lead Number']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font = Font(bold=True, size=10, name="Inter", color=TEXT_DARK)  # Bold like HTML template
                    
                elif col_name in ['Disposition', 'Lead Category', 'Engagement Level']:
                    cell.alignment = center_alignment
                    cell.font = Font(size=10, name="Segoe UI", color=TEXT_DARK)
                    
                elif col_name in ['Summary', 'Key Discussion Points', 'Recommendations', 'Prospect Questions', 'Prospect Concerns', 'Recommended Action']:
                    cell.alignment = data_alignment
                    cell.font = Font(size=9, name="Segoe UI", color=TEXT_DARK)
                    
                else:
                    cell.alignment = center_alignment
                    cell.font = Font(size=10, name="Segoe UI", color=TEXT_DARK)
            
        # Executive table row spacing (matching HTML template)
        worksheet.row_dimensions[row_idx].height = 28
        
    # Executive column widths (optimized for professional dashboard view)
    column_widths = {
        'Call Log ID': 16,
        'Lead Name': 20,
        'Lead Number': 16,
        'Date': 14,
        'Start Time (GST)': 16,
        'End Time (GST)': 16,
        'Duration': 12,
        'Call Status': 15,
        'Time Period (GST)': 20,
        'Disposition': 18,
        'Recommended Action': 40,
        'Lead Category': 16,
        'Engagement Level': 16,
        'Sentiment': 22,
        'Summary': 50,
        'Key Discussion Points': 40,
        'Prospect Questions': 40,
        'Prospect Concerns': 40,
        'Recommendations': 40,
    }
    
    # Apply executive dashboard column widths
    for idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        
        if col in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col]
        else:
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 3, 50)
    
    # Executive navigation (freeze panes at data start)
    worksheet.freeze_panes = f'A{header_row + 1}'
    
    # Professional filtering system
    worksheet.auto_filter.ref = f'A{header_row}:{get_column_letter(len(df.columns))}{worksheet.max_row}'


def export_to_excel(
    data: List[Dict],
    batch_id: str,
    batch_info: Optional[Dict] = None,
    output_file: Optional[str] = None
) -> str:
    """
    Export batch call data to Excel file
    
    Args:
        data: List of call data dictionaries
        batch_id: Batch UUID for filename
        batch_info: Optional batch metadata
        output_file: Optional custom output path
    
    Returns:
        Path to exported Excel file
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("pandas/openpyxl not installed. Install with: pip install pandas openpyxl")
    
    if not data:
        raise ValueError("No data to export")
    
    # Generate output filename
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        short_id = str(batch_id)[:8]
        output_file = f"batch_report_{short_id}_{timestamp}.xlsx"
    
    output_path = EXPORTS_DIR / output_file
    
    # Prepare data for Excel
    excel_data = []
    for record in data:
        started_at = record.get('started_at')
        ended_at = record.get('ended_at')
        
        # Convert timestamps to Gulf Standard Time (UTC+4)
        gst_started_at = _convert_to_gst(started_at) if isinstance(started_at, datetime) else None
        gst_ended_at = _convert_to_gst(ended_at) if isinstance(ended_at, datetime) else None
        
        # Format timestamps in GST
        date_str = gst_started_at.strftime('%Y-%m-%d') if gst_started_at else 'N/A'
        start_time_str = gst_started_at.strftime('%H:%M:%S') if gst_started_at else 'N/A'
        end_time_str = gst_ended_at.strftime('%H:%M:%S') if gst_ended_at else 'N/A'
        
        row = {
            'Call Log ID': str(record.get('call_log_id', '')) if record.get('call_log_id') else '',
            'Lead Name': record.get('lead_name', '') or '',
            'Lead Number': record.get('lead_number', '') or '',
            'Date': date_str,
            'Start Time (GST)': start_time_str,
            'End Time (GST)': end_time_str,
            'Duration': _format_duration(started_at, ended_at),
            'Call Status': record.get('call_status', '') or '',
            'Time Period (GST)': _get_time_period(started_at),
            'Disposition': record.get('disposition', '') or '',
            'Recommended Action': record.get('recommended_action', '') or '',
            'Lead Category': record.get('lead_category', '') or '',
            'Engagement Level': record.get('engagement_level', '') or '',
            'Sentiment': record.get('sentiment_description', '') or '',
            'Summary': record.get('summary', '') or '',
            'Key Discussion Points': _format_list_field(record.get('key_discussion_points')),
            'Prospect Questions': _format_list_field(record.get('prospect_questions')),
            'Prospect Concerns': _format_list_field(record.get('prospect_concerns')),
            'Recommendations': record.get('recommendations', '') or '',
        }
        excel_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Column order
    column_order = [
        'Call Log ID', 'Lead Name', 'Lead Number', 'Date', 'Start Time (GST)',
        'End Time (GST)', 'Duration', 'Call Status', 'Time Period (GST)', 'Disposition',
        'Recommended Action', 'Lead Category', 'Engagement Level', 'Sentiment',
        'Summary', 'Key Discussion Points', 'Prospect Questions',
        'Prospect Concerns', 'Recommendations'
    ]
    
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary sheet
        if batch_info:
            summary_data = {
                'Metric': ['Batch ID', 'Status', 'Total Calls', 'Completed', 'Failed', 
                          'Cancelled', 'Created At', 'Completed At'],
                'Value': [
                    batch_info.get('id', 'N/A'),
                    batch_info.get('status', 'N/A'),
                    batch_info.get('total_calls', 0),
                    batch_info.get('completed_calls', 0),
                    batch_info.get('failed_calls', 0),
                    batch_info.get('cancelled_calls', 0),
                    str(batch_info.get('created_at', 'N/A')),
                    str(batch_info.get('completed_at', 'N/A')),
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            try:
                # Apply the same enhanced executive formatting used for other sheets
                _apply_enhanced_excel_formatting(writer.sheets['Summary'], summary_df)
            except Exception:
                # Don't fail the whole export if summary formatting has an issue
                logger.exception("Failed to apply enhanced formatting to Summary sheet")
        
        # Group by time period
        if len(data) > 1 and 'Time Period (GST)' in df.columns:
            time_periods = df['Time Period (GST)'].unique()
            for period in sorted(time_periods):
                period_df = df[df['Time Period (GST)'] == period].copy()
                period_df = period_df.drop(columns=['Time Period (GST)'], errors='ignore')
                
                sheet_name = period[:31].replace(':', '-').replace('/', '-')
                period_df.to_excel(writer, sheet_name=sheet_name, index=False)
                _apply_enhanced_excel_formatting(writer.sheets[sheet_name], period_df)
            
            # Also add an 'All Calls' sheet
            df_all = df.drop(columns=['Time Period (GST)'], errors='ignore')
            df_all.to_excel(writer, sheet_name='All Calls', index=False)
            _apply_enhanced_excel_formatting(writer.sheets['All Calls'], df_all)
        else:
            df.to_excel(writer, sheet_name='Call Analysis', index=False)
            _apply_enhanced_excel_formatting(writer.sheets['Call Analysis'], df)
    
    logger.info(f"Excel report created: {output_path} ({len(data)} records)")
    return str(output_path)


def _get_email_method() -> str:
    """
    Get the configured email method.
    
    Returns:
        "smtp" (default) or "oauth"
    """
    method = os.getenv("BATCH_EMAIL_METHOD", "smtp").lower().strip()
    if method not in ("smtp", "oauth"):
        logger.warning(f"Invalid BATCH_EMAIL_METHOD '{method}', defaulting to 'smtp'")
        return "smtp"
    return method


async def _send_email_via_oauth(
    to_emails: List[str],
    subject: str,
    html_body: str,
    text_body: str,
    attachment_path: Optional[str] = None,
    oauth_user_id: Optional[int] = None,
) -> bool:
    """
    Send email using Google OAuth tokens via Gmail API.
    
    Args:
        to_emails: List of recipient email addresses
        subject: Email subject
        html_body: HTML email body
        text_body: Plain text email body (fallback)
        attachment_path: Optional path to file attachment
        oauth_user_id: User ID for OAuth tokens (defaults to BATCH_EMAIL_OAUTH_USER_ID env var)
    
    Returns:
        True if email sent successfully
    """
    if not OAUTH_EMAIL_AVAILABLE:
        logger.error("OAuth email not available (missing dependencies)")
        return False
    
    # Get the user ID for OAuth tokens - use provided value or fall back to env var
    if oauth_user_id is not None:
        oauth_user_id_int = oauth_user_id
        logger.info(f"Using provided oauth_user_id={oauth_user_id_int} (from batch initiated_by)")
    else:
        oauth_user_id_str = os.getenv("BATCH_EMAIL_OAUTH_USER_ID", "10")
        try:
            oauth_user_id_int = int(oauth_user_id_str)
            logger.info(f"Using fallback BATCH_EMAIL_OAUTH_USER_ID={oauth_user_id_int} from env var")
        except ValueError:
            logger.error(f"Invalid BATCH_EMAIL_OAUTH_USER_ID: {oauth_user_id_str}")
            return False
    
    logger.info(f"Sending email via OAuth using user_id={oauth_user_id_int}")
    
    try:
        # Resolve credentials using existing infrastructure
        resolver = GoogleCredentialResolver()
        credentials = await resolver.load_credentials(oauth_user_id_int)
        
        # Ensure token is valid
        if credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
        
        # Build the email message
        msg = MIMEMultipart("mixed")
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = subject
        
        # Create alternative part for text/html
        alt_part = MIMEMultipart("alternative")
        alt_part.attach(MIMEText(text_body, "plain"))
        alt_part.attach(MIMEText(html_body, "html"))
        msg.attach(alt_part)
        
        # Add attachment if provided
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename={os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        
        # Encode and send via Gmail API
        encoded_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        
        service = build("gmail", "v1", credentials=credentials, cache_discovery=False)
        response = service.users().messages().send(
            userId="me",
            body={"raw": encoded_message}
        ).execute()
        
        logger.info(f"OAuth email sent successfully to {len(to_emails)} recipient(s), message_id={response.get('id')}")
        return True
        
    except GoogleCredentialError as e:
        logger.error(f"Failed to load OAuth credentials for user {oauth_user_id}: {e}")
        return False
    except Exception as e:
        logger.error(f"Failed to send OAuth email: {e}", exc_info=True)
        return False


def _send_email_via_smtp(
    to_emails: List[str],
    subject: str,
    html_body: str,
    text_body: str,
    attachment_path: Optional[str] = None,
) -> bool:
    """
    Send email using SMTP.
    
    Args:
        to_emails: List of recipient email addresses
        subject: Email subject
        html_body: HTML email body
        text_body: Plain text email body (fallback)
        attachment_path: Optional path to file attachment
    
    Returns:
        True if email sent successfully
    """
    if not EMAIL_AVAILABLE:
        logger.error("SMTP email libraries not available")
        return False
    
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("FROM_EMAIL") or smtp_user
    
    if not smtp_user or not smtp_password:
        logger.warning("SMTP credentials not configured. Set SMTP_USER and SMTP_PASSWORD in .env")
        return False
    
    if not to_emails:
        logger.warning("No email recipients specified")
        return False
    
    if attachment_path and not os.path.exists(attachment_path):
        logger.error(f"Attachment file not found: {attachment_path}")
        return False
    
    try:
        msg = MIMEMultipart("mixed")
        msg['From'] = from_email
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = subject
        
        # Create alternative part for text/html
        alt_part = MIMEMultipart("alternative")
        alt_part.attach(MIMEText(text_body, "plain"))
        alt_part.attach(MIMEText(html_body, "html"))
        msg.attach(alt_part)
        
        if attachment_path:
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename={os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        
        server.sendmail(from_email, to_emails, msg.as_string())
        server.quit()
        
        logger.info(f"SMTP email sent successfully to {len(to_emails)} recipient(s)")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send SMTP email: {e}", exc_info=True)
        return False


async def send_report_email(
    to_emails: List[str],
    subject: str,
    body: str,
    attachment_path: Optional[str] = None,
    batch_info: Optional[Dict] = None,
    calls_with_analysis: int = 0,
    calls_without_analysis: int = 0,
    excel_path: Optional[str] = None,
    oauth_user_id: Optional[int] = None,
) -> bool:
    """
    Send email with optional attachment using configured method (SMTP or OAuth).
    
    The email method is determined by BATCH_EMAIL_METHOD environment variable:
    - "smtp" (default): Use traditional SMTP
    - "oauth": Use Google Gmail API with OAuth tokens
    
    Args:
        to_emails: List of recipient email addresses
        subject: Email subject
        body: Email body text (plain text fallback)
        attachment_path: Optional path to file attachment
        batch_info: Optional batch metadata for HTML email
        calls_with_analysis: Number of calls with analysis
        calls_without_analysis: Number of calls without analysis
        excel_path: Path to Excel file for HTML email
        oauth_user_id: User ID for OAuth tokens (used when BATCH_EMAIL_METHOD=oauth)
    
    Returns:
        True if email sent successfully
    """
    if not to_emails:
        logger.warning("No email recipients specified")
        return False
    
    # Build HTML email body (enhanced formatting)
    html_body = _build_html_email_body(
        batch_info=batch_info,
        calls_with_analysis=calls_with_analysis,
        calls_without_analysis=calls_without_analysis,
        has_excel=excel_path is not None,
    )
    
    email_method = _get_email_method()
    logger.info(f"Sending batch report email via {email_method.upper()} to {len(to_emails)} recipient(s)")
    
    if email_method == "oauth":
        return await _send_email_via_oauth(
            to_emails=to_emails,
            subject=subject,
            html_body=html_body,
            text_body=body,
            attachment_path=attachment_path,
            oauth_user_id=oauth_user_id,
        )
    else:
        return _send_email_via_smtp(
            to_emails=to_emails,
            subject=subject,
            html_body=html_body,
            text_body=body,
            attachment_path=attachment_path,
        )


def _build_html_email_body(
    batch_info: Optional[Dict] = None,
    calls_with_analysis: int = 0,
    calls_without_analysis: int = 0,
    has_excel: bool = True,
) -> str:
    """
    Build a professional HTML email body for batch reports.
    
    Args:
        batch_info: Batch metadata dictionary
        calls_with_analysis: Number of calls with analysis
        calls_without_analysis: Number of calls without analysis  
        has_excel: Whether Excel report is attached
    
    Returns:
        HTML email body string
    """
    if batch_info is None:
        batch_info = {}
    
    # Get current time in GST
    gst_now = datetime.now(GST)
    report_time = gst_now.strftime('%Y-%m-%d %H:%M:%S')
    
    # Calculate statistics
    total_calls = batch_info.get('total_calls', 0)
    completed_calls = batch_info.get('completed_calls', 0)
    failed_calls = batch_info.get('failed_calls', 0)
    cancelled_calls = batch_info.get('cancelled_calls', 0)
    
    # Status color
    status = batch_info.get('status', 'unknown')
    status_colors = {
        'completed': '#28a745',  # Green
        'stopped': '#ffc107',    # Yellow
        'cancelled': '#dc3545',  # Red
        'running': '#17a2b8',    # Blue
        'pending': '#6c757d',    # Gray
    }
    status_color = status_colors.get(status.lower(), '#6c757d')
    
    # Success rate calculation
    success_rate = round((completed_calls / total_calls * 100), 1) if total_calls > 0 else 0
    analysis_rate = round((calls_with_analysis / total_calls * 100), 1) if total_calls > 0 else 0
    
    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Strategic Call Analytics Report</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    </style>
</head>
<body style="margin: 0; padding: 0; font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f8f9fa; line-height: 1.6;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background: #f8f9fa; padding: 30px 20px;">
        <tr>
            <td>
                <table role="presentation" width="800" cellspacing="0" cellpadding="0" border="0" align="center" style="background: #ffffff; box-shadow: 0 4px 6px rgba(0,0,0,0.02), 0 12px 24px rgba(0,0,0,0.08); border-radius: 12px; overflow: hidden;">
                    
                    <tr>
                        <td>
                            </head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background-color: #f4f4f4;">
        <tr>
            <td style="padding: 20px 0;">
                <table role="presentation" width="600" cellspacing="0" cellpadding="0" border="0" align="center" style="background-color: #ffffff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    
                    <!-- Header -->
                    <tr>
                        <td style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px 40px; border-radius: 8px 8px 0 0;">
                            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 600;">
                                📊 Batch Call Report
                            </h1>
                            <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0 0; font-size: 14px;">
                                Automated Voice Agent Report
                            </p>
                        </td>
                    </tr>
                    </td>
                    </tr>
                    
                    <!-- Status Badge -->
                    <tr>
                        <td style="padding: 25px 40px 15px 40px;">
                            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
                                <tr>
                                    <td>
                                        <span style="display: inline-block; background-color: {status_color}; color: #ffffff; padding: 8px 20px; border-radius: 20px; font-size: 14px; font-weight: 600; text-transform: uppercase;">
                                            {status.title()}
                                        </span>
                                    </td>
                                    <td style="text-align: right; color: #666666; font-size: 13px;">
                                        Generated: {report_time} GST
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Batch ID -->
                    <tr>
                        <td style="padding: 0 40px 20px 40px;">
                            <p style="margin: 0; color: #888888; font-size: 12px;">BATCH ID</p>
                            <p style="margin: 5px 0 0 0; color: #333333; font-size: 14px; font-family: 'Courier New', monospace; word-break: break-all;">
                                {batch_info.get('id', 'N/A')}
                            </p>
                        </td>
                    </tr>
                    
                    <!-- Stats Grid -->
                    <tr>
                        <td style="padding: 0 40px 30px 40px;">
                            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
                                <tr>
                                    <!-- Total Calls -->
                                    <td width="25%" style="text-align: center; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
                                        <p style="margin: 0; font-size: 32px; font-weight: 700; color: #333333;">{total_calls}</p>
                                        <p style="margin: 5px 0 0 0; font-size: 11px; color: #666666; text-transform: uppercase; letter-spacing: 0.5px;">Total Calls</p>
                                    </td>
                                    <td width="2%"></td>
                                    <!-- Completed -->
                                    <td width="23%" style="text-align: center; padding: 15px; background-color: #d4edda; border-radius: 8px;">
                                        <p style="margin: 0; font-size: 32px; font-weight: 700; color: #28a745;">{completed_calls}</p>
                                        <p style="margin: 5px 0 0 0; font-size: 11px; color: #155724; text-transform: uppercase; letter-spacing: 0.5px;">Completed</p>
                                    </td>
                                    <td width="2%"></td>
                                    <!-- Failed -->
                                    <td width="23%" style="text-align: center; padding: 15px; background-color: #f8d7da; border-radius: 8px;">
                                        <p style="margin: 0; font-size: 32px; font-weight: 700; color: #dc3545;">{failed_calls}</p>
                                        <p style="margin: 5px 0 0 0; font-size: 11px; color: #721c24; text-transform: uppercase; letter-spacing: 0.5px;">Failed</p>
                                    </td>
                                    <td width="2%"></td>
                                    <!-- Cancelled -->
                                    <td width="23%" style="text-align: center; padding: 15px; background-color: #fff3cd; border-radius: 8px;">
                                        <p style="margin: 0; font-size: 32px; font-weight: 700; color: #856404;">{cancelled_calls}</p>
                                        <p style="margin: 5px 0 0 0; font-size: 11px; color: #856404; text-transform: uppercase; letter-spacing: 0.5px;">Cancelled</p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Progress Bars -->
                    <tr>
                        <td style="padding: 0 40px 30px 40px;">
                            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
                                <!-- Success Rate -->
                                <tr>
                                    <td style="padding-bottom: 15px;">
                                        <p style="margin: 0 0 8px 0; font-size: 13px; color: #333333;">
                                            <strong>Success Rate:</strong> {success_rate}%
                                        </p>
                                        <div style="background-color: #e9ecef; border-radius: 10px; height: 10px; overflow: hidden;">
                                            <div style="background-color: #28a745; height: 100%; width: {success_rate}%; border-radius: 10px;"></div>
                                        </div>
                                    </td>
                                </tr>
                                <!-- Analysis Rate -->
                                <tr>
                                    <td>
                                        <p style="margin: 0 0 8px 0; font-size: 13px; color: #333333;">
                                            <strong>Analysis Coverage:</strong> {analysis_rate}% ({calls_with_analysis} of {total_calls} calls)
                                        </p>
                                        <div style="background-color: #e9ecef; border-radius: 10px; height: 10px; overflow: hidden;">
                                            <div style="background-color: #667eea; height: 100%; width: {analysis_rate}%; border-radius: 10px;"></div>
                                        </div>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Timestamps -->
                    <tr>
                        <td style="padding: 0 40px 30px 40px;">
                            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background-color: #f8f9fa; border-radius: 8px; padding: 15px;">
                                <tr>
                                    <td style="padding: 15px;">
                                        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
                                            <tr>
                                                <td width="50%">
                                                    <p style="margin: 0; font-size: 12px; color: #666666;">Started</p>
                                                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #333333;">
                                                        {str(batch_info.get('created_at', 'N/A'))[:19]}
                                                    </p>
                                                </td>
                                                <td width="50%">
                                                    <p style="margin: 0; font-size: 12px; color: #666666;">Completed</p>
                                                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #333333;">
                                                        {str(batch_info.get('completed_at', 'N/A'))[:19] if batch_info.get('completed_at') else 'N/A'}
                                                    </p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Attachment Notice -->
                    <tr>
                        <td style="padding: 0 40px 30px 40px;">
                            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background-color: #e7f1ff; border-left: 4px solid #667eea; border-radius: 4px;">
                                <tr>
                                    <td style="padding: 15px 20px;">
                                        <p style="margin: 0; font-size: 14px; color: #333333;">
                                            📎 {"<strong>Excel report attached</strong> - Contains detailed call analysis organized by time periods (GST timezone)." if has_excel else "<em>Note: Excel report could not be generated (missing dependencies).</em>"}
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f8f9fa; padding: 25px 40px; border-radius: 0 0 8px 8px; text-align: center;">
                            <p style="margin: 0; font-size: 12px; color: #888888;">
                                This is an automated report from the Voice Agent System.
                            </p>
                            <p style="margin: 10px 0 0 0; font-size: 11px; color: #aaaaaa;">
                                © {gst_now.year} Dev-S-t • All times shown in Gulf Standard Time (GST/UTC+4)
                            </p>
                        </td>
                    </tr>
                    
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
"""
    return html


async def generate_batch_report(
    batch_id: str,
    send_email: bool = False,
    email_recipients: Optional[List[str]] = None,
    wait_for_analysis: bool = True,
    max_wait_seconds: int = 120,
) -> Dict[str, Any]:
    """
    Generate a batch report with Excel export and optional email
    
    This is the main entry point for integration with main.py
    Call this when a batch completes/stops/cancels.
    
    Args:
        batch_id: Batch UUID
        send_email: Whether to send email with report
        email_recipients: Optional list of email addresses (defaults to env var)
        wait_for_analysis: Whether to wait for post-call analysis to complete (default True)
        max_wait_seconds: Maximum time to wait for analysis (default 120s)
    
    Returns:
        Dict with status, file_path, email_sent, etc.
    """
    logger.info(f"Generating batch report for batch_id={batch_id}")
    
    # Get batch info
    batch_info = await get_batch_info(batch_id)
    if not batch_info:
        logger.error(f"Batch {batch_id} not found")
        return {
            "status": "error",
            "message": f"Batch {batch_id} not found",
            "file_path": None,
            "email_sent": False,
            "call_count": 0,
        }
    
    logger.info(f"Batch found: status={batch_info['status']}, total={batch_info['total_calls']}")
    
    # Wait for post-call analysis to complete before fetching data
    if wait_for_analysis:
        analysis_counts = await wait_for_analysis_completion(
            batch_id, 
            max_wait_seconds=max_wait_seconds,
            poll_interval_seconds=10
        )
        logger.info(f"Analysis complete: {analysis_counts['analyzed_calls']}/{analysis_counts['total_calls']} calls analyzed")
    
    # Fetch call data (now with analysis data populated)
    call_data = await fetch_batch_call_data(batch_id)
    
    calls_with_analysis = len([c for c in call_data if c.get('summary')])
    calls_without_analysis = len(call_data) - calls_with_analysis
    
    logger.info(f"Fetched {len(call_data)} calls ({calls_with_analysis} with analysis)")
    
    if not call_data:
        logger.warning(f"No calls found for batch {batch_id}")
        return {
            "status": "warning",
            "message": f"No calls found for batch {batch_id}",
            "file_path": None,
            "email_sent": False,
            "call_count": 0,
        }
    
    # Export to Excel
    excel_path = None
    try:
        excel_path = export_to_excel(call_data, batch_id, batch_info)
        logger.info(f"Excel report created: {excel_path}")
    except ImportError as e:
        logger.warning(f"Excel export not available: {e}")
        # Continue without Excel - we can still send a text-only email
    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}", exc_info=True)
        return {
            "status": "error",
            "message": f"Failed to create Excel report: {e}",
            "file_path": None,
            "email_sent": False,
            "call_count": len(call_data),
        }
    
    # Get current time in GST for report
    gst_now = datetime.now(GST)
    
    # Send email if requested
    email_sent = False
    if send_email:
        recipients = email_recipients or await get_email_recipients_async(
            initiated_by=batch_info.get('initiated_by')
        )
        if recipients:
            subject = f"Strategic Call Analytics Report – {batch_info['status'].title()} – {len(call_data)} Voice Interactions"
            body = f"""
STRATEGIC CALL ANALYTICS REPORT
Executive Summary of Voice Agent Performance

═══════════════════════════════════════════
BATCH OVERVIEW
═══════════════════════════════════════════
Batch ID: {batch_id}
Status: {batch_info['status'].upper()}
Total Voice Interactions: {batch_info['total_calls']}
Successfully Completed: {batch_info['completed_calls']}
Failed Attempts: {batch_info['failed_calls']}
Cancelled Operations: {batch_info['cancelled_calls']}

═══════════════════════════════════════════
ANALYTICS COVERAGE
═══════════════════════════════════════════
Calls with AI Analysis: {calls_with_analysis}
Pending Analysis: {calls_without_analysis}
Analysis Coverage: {round((calls_with_analysis / len(call_data) * 100), 1) if call_data else 0}%

═══════════════════════════════════════════
REPORT DETAILS
═══════════════════════════════════════════
Generated: {gst_now.strftime('%Y-%m-%d %H:%M:%S')} GST (Gulf Standard Time)
Report Type: Comprehensive Executive Dashboard

{"📊 The attached Excel report provides detailed performance analytics organized by time periods (GST) with executive-grade formatting and interactive filtering." if excel_path else "⚠️  Note: Excel report could not be generated (missing dependencies)."}

═══════════════════════════════════════════

This automated report is generated by the Voice Agent Analytics Platform.
For questions or technical support, please contact your system administrator.

            """
            
            email_sent = await send_report_email(
                to_emails=recipients,
                subject=subject,
                body=body,
                attachment_path=excel_path,
                batch_info=batch_info,
                calls_with_analysis=calls_with_analysis,
                calls_without_analysis=calls_without_analysis,
                excel_path=excel_path,
                oauth_user_id=batch_info.get('initiated_by'),  # Use batch initiator for OAuth email
            )
        else:
            logger.info("No email recipients configured - skipping email")
    
    return {
        "status": "success",
        "message": "Batch report generated successfully",
        "file_path": excel_path,
        "email_sent": email_sent,
        "call_count": len(call_data),
        "calls_with_analysis": calls_with_analysis,
        "calls_without_analysis": calls_without_analysis,
        "batch_info": batch_info,
    }


async def print_batch_list():
    """Print list of available batches"""
    batches = await list_batches(50)
    
    if not batches:
        print("No batches found in database")
        return
    
    print(f"\n{'='*100}")
    print("AVAILABLE BATCHES (Last 50)")
    print(f"{'='*100}")
    print(f"{'Batch ID':<40} {'Status':<12} {'Total':<8} {'Completed':<10} {'Failed':<8} {'Created'}")
    print("-" * 100)
    
    for batch in batches:
        created = batch['created_at'].strftime('%Y-%m-%d %H:%M') if batch['created_at'] else 'N/A'
        print(f"{batch['id']:<40} {batch['status']:<12} {batch['total_calls']:<8} "
              f"{batch['completed_calls']:<10} {batch['failed_calls']:<8} {created}")
    
    print("-" * 100)
    print(f"\nTo generate a report, use:")
    print(f"  uv run python -m post_call_analysis.batch_report_service --batch-id <BATCH_ID>")


def main():
    """CLI entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Batch Call Report Service',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # List all batches
    uv run python -m post_call_analysis.batch_report_service --list-batches
    
    # Generate report for a batch
    uv run python -m post_call_analysis.batch_report_service --batch-id <UUID>
    
    # Generate and email report
    uv run python -m post_call_analysis.batch_report_service --batch-id <UUID> --send-email
        """
    )
    
    parser.add_argument('--list-batches', action='store_true', help='List available batches')
    parser.add_argument('--batch-id', help='Batch UUID to generate report for')
    parser.add_argument('--send-email', action='store_true', help='Send report via email')
    parser.add_argument('--email', help='Override email recipient')
    
    args = parser.parse_args()
    
    if args.list_batches:
        asyncio.run(print_batch_list())
        return 0
    
    if not args.batch_id:
        parser.error("Either --list-batches or --batch-id is required")
    
    email_recipients = [args.email] if args.email else None
    
    result = asyncio.run(generate_batch_report(
        args.batch_id,
        send_email=args.send_email,
        email_recipients=email_recipients,
    ))
    
    print(f"\n{'='*60}")
    print("REPORT SUMMARY")
    print(f"{'='*60}")
    print(f"Status: {result['status']}")
    print(f"Message: {result['message']}")
    print(f"File Path: {result.get('file_path', 'N/A')}")
    print(f"Email Sent: {'Yes' if result.get('email_sent') else 'No'}")
    print(f"Call Count: {result.get('call_count', 0)}")
    print(f"With Analysis: {result.get('calls_with_analysis', 0)}")
    print(f"Without Analysis: {result.get('calls_without_analysis', 0)}")
    print(f"{'='*60}\n")
    
    return 0 if result['status'] == 'success' else 1


if __name__ == "__main__":
    exit(main())
